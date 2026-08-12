# ============================================================
# Animal Supplement Optimizer v3 (Streamlit App)
#
# Changes vs app_2:
# - Patient selection follows the full hierarchy:
#     Animal → Fall → Diagnose → extra inputs
# - Min values read directly from Bedarf sheet L–BK columns
#   via xlwings (Excel recalculates with user-supplied inputs)
# - Max = min × 3 (Vit.A/D3/Cu/Zn/Jod/Selen) or min × 5
# - Original Excel is never modified (temp-copy approach)
# ============================================================

import base64
import hashlib
import io
import math
import os
import re
import shutil
import tempfile
import warnings
import difflib
import unicodedata
from typing import Optional

import plotly.graph_objects as go

import msoffcrypto
import numpy as np
import openpyxl
import pandas as pd
import pulp
import streamlit as st
import streamlit.components.v1 as _st_components

import platform as _platform
try:
    import xlwings as xw
    _XLWINGS_OK = _platform.system() == "Windows"
except ImportError:
    _XLWINGS_OK = False

# ------------------------------------------------------------
# 0) PAGE CONFIG
# ------------------------------------------------------------
st.set_page_config(page_title="Animal Nutrition Optimizer", layout="wide")

# ------------------------------------------------------------
# 0b) GLOBAL CSS
# ------------------------------------------------------------
st.markdown(
    """
    <style>
    .block-container{padding-top:1.5rem!important;}
    section[data-testid="stMain"] .block-container{padding-top:1.5rem!important;}
    div[data-testid="column"]:last-child div[data-testid="stImage"]{display:flex;justify-content:flex-end;}
    div[data-testid="column"]:last-child div[data-testid="stImage"] img{margin-left:auto;}
    :root{
      --btn_primary_font: 1.15rem; --btn_primary_weight: 600;
      --btn_primary_py: 0.40rem;  --btn_primary_px: 0.75rem;
      --btn_primary_radius: 10px;
      --btn_secondary_font: 1rem; --btn_secondary_weight: 450;
      --btn_secondary_py: 0.20rem;   --btn_secondary_px: 0.55rem;
      --btn_secondary_radius: 10px;
      --chk_font: 1rem; --chk_weight: 500;
      --chk_scale: 1.1;   --chk_gap: 6px;
      --caption_font: 0.95rem; --caption_weight: 450;
      --caption_opacity: 0.85; --caption_line: 1.35;
    }
    div[data-testid="stCheckbox"] label,
    div[data-testid="stCheckbox"] label p,
    div[data-testid="stCheckbox"] label span{
      font-size:var(--chk_font)!important; font-weight:var(--chk_weight)!important;
      line-height:1.35!important;
    }
    div[data-testid="stCheckbox"] input{
      transform:scale(var(--chk_scale))!important; margin-right:var(--chk_gap)!important;
    }
    div[data-testid="stCaption"],div[data-testid="stCaption"] *{
      font-size:var(--caption_font)!important; font-weight:var(--caption_weight)!important;
      line-height:var(--caption_line)!important; opacity:var(--caption_opacity)!important;
    }
    div[data-testid="stMarkdownContainer"] small,
    div[data-testid="stMarkdownContainer"] small *{
      font-size:var(--caption_font)!important; font-weight:var(--caption_weight)!important;
      line-height:var(--caption_line)!important; opacity:var(--caption_opacity)!important;
    }
    .caption-note{font-size:var(--caption_font);font-weight:var(--caption_weight);
      line-height:var(--caption_line);opacity:var(--caption_opacity);
      margin-top:0.15rem;margin-bottom:0.40rem;}
    button[kind="primary"],button[data-testid="baseButton-primary"]{
      padding:var(--btn_primary_py) var(--btn_primary_px)!important;
      border-radius:var(--btn_primary_radius)!important;}
    button[kind="primary"] *,button[data-testid="baseButton-primary"] *{
      font-size:var(--btn_primary_font)!important;font-weight:var(--btn_primary_weight)!important;
      line-height:1.1!important;}
    button[kind="secondary"],button[data-testid="baseButton-secondary"]{
      padding:var(--btn_secondary_py) var(--btn_secondary_px)!important;
      border-radius:var(--btn_secondary_radius)!important;}
    button[kind="secondary"] *,button[data-testid="baseButton-secondary"] *{
      font-size:var(--btn_secondary_font)!important;font-weight:var(--btn_secondary_weight)!important;
      line-height:1.2!important;}
    .upload-title{font-size:1.10rem;font-weight:750;margin-bottom:0.1rem;}
    .muted-hint{color:rgba(49,51,63,0.7);font-size:1.05rem;margin-top:-0.1rem;margin-bottom:0.6rem;}
    .statusline{font-size:1.05rem;font-weight:650;margin:0.15rem 0 0.65rem 0;}
    .okrow,.warnrow,.errorrow{border-radius:10px;padding:0.55rem 0.7rem;
      margin:0.25rem 0 0.25rem 0;font-weight:650;}
    .okrow{background:rgba(46,184,92,0.12);border:1px solid rgba(46,184,92,0.30);}
    .warnrow{background:rgba(255,165,0,0.12);border:1px solid rgba(255,165,0,0.30);}
    .errorrow{background:rgba(255,0,0,0.08);border:1px solid rgba(255,0,0,0.18);}
    .warnrow-after{margin-bottom:0.95rem!important;}
    .nutrient-actions{margin-top:0rem;}
    div[data-testid="stDataEditor"],div[data-testid="stDataFrame"]{
      margin-bottom:0.10rem!important;padding-bottom:0rem!important;}
    .thin-sep{border:0;border-top:1px solid rgba(49,51,63,0.22);
      margin:0.20rem 0 0.20rem 0;padding:0;}
    .kpi-card{border-radius:12px;padding:16px 18px;
      border:1px solid rgba(49,51,63,0.12);background:rgba(49,51,63,0.05);}
    .kpi-title{font-size:1.05rem;font-weight:700;opacity:0.85;}
    .kpi-value{font-size:2.2rem;font-weight:900;margin-top:4px;line-height:1.05;}
    div[data-testid="stVerticalBlock"]:has(#ft-sm-btn-marker) button[kind="secondary"]{
      font-size:0.85rem!important;padding:0.1rem 0.4rem!important;min-height:auto!important;}
    </style>
    """,
    unsafe_allow_html=True,
)

# Scroll-position preservation across Streamlit reruns.
# st.markdown <script> tags are never executed by browsers (innerHTML restriction);
# components.html runs in an iframe where JS executes, and window.parent gives
# access to the main page scroll.
_st_components.html("""
<script>
(function(){
  var p = window.parent, _y = 0, _t = 0, _tmr;
  p.document.addEventListener('click', function(){ _y = p.scrollY; _t = Date.now(); }, true);
  new MutationObserver(function(){
    clearTimeout(_tmr);
    _tmr = setTimeout(function(){
      if(p.scrollY < 60 && _y > 80 && Date.now()-_t < 3000) p.scrollTo(0, _y);
    }, 100);
  }).observe(p.document.body, {childList:true, subtree:false});
})();
</script>
""", height=0)

# ------------------------------------------------------------
# 1) HEADER
# ------------------------------------------------------------
#st.title("🐾💊 Animal Nutrition Optimizer feat. CarniDiet©")
#st.image("static/Vetmedlogo.png", width="stretch")

col_left, col_right = st.columns([4, 1.5])
with col_left:
    #st.title("🐾💊 Animal Nutrition Optimizer feat. CarniDiet©")
    try:
        _bowl_b64 = base64.b64encode(open("static/dog-bowl.svg", "rb").read()).decode()
        _bowl_img = f"<img src='data:image/svg+xml;base64,{_bowl_b64}' style='height:2.8rem;vertical-align:middle;margin-right:0.3rem;margin-left:-0.5rem;'>"
        _catdog_b64 = base64.b64encode(open("static/cat-dog.svg", "rb").read()).decode()
        _catdog_img = f"<img src='data:image/svg+xml;base64,{_catdog_b64}' style='height:2.8rem;vertical-align:middle;margin-right:0.3rem;position:relative;top:-0.6rem;'>"
    except Exception:
        _bowl_img = ""
        _catdog_img = ""
    st.markdown(
        f'<h1 style="font-size:2.5rem;font-weight:800;margin:0;color:#19425e;">{_catdog_img}{_bowl_img} Animal Nutrition Optimizer '
        '<span style="font-size:1.25rem;font-weight:400;color:#3a7fc1;vertical-align:baseline;">'
        'feat. CarniDiet©</span></h1>',
        unsafe_allow_html=True,
    )
with col_right:
    try:
        _logo_b64 = base64.b64encode(open("static/Vetmedlogo.png", "rb").read()).decode()
        st.markdown(
            f"<div style='text-align:right;margin-top:1.3rem;'>"
            f"<img src='data:image/png;base64,{_logo_b64}' width='120'>"
            f"</div>",
            unsafe_allow_html=True,
        )
    except Exception:
        pass
    st.markdown(
        "<div style='text-align:right;font-size:0.75rem;line-height:1.3;'>Univ.-Prof. Dr. Qendrim Zebeli</div>",
        unsafe_allow_html=True,
    )
st.markdown(
    "<div style='text-align:left;'><div style='background:rgba(28,131,225,0.12);border:1px solid rgba(28,131,225,0.35);"
    "border-radius:0.5rem;padding:0.75rem 1rem;font-size:0.95rem;line-height:1.4;color:#5385a6;font-weight:600;display:inline-block;text-align:left;'>"
    #"ℹ️ CarniDiet© ist ausschließlich zu Übungszwecken für Studierende der Vetmeduni Vienna genehmigt. "
    "<span style='line-height:0;display:block;margin-top:5px;margin-bottom:0.6rem;'>ℹ️ CarniDiet© ist ausschließlich zu Übungszwecken für Studierende der Vetmeduni Vienna genehmigt.</span>"
    "Die Weitergabe, kommerz. Nutzung und Vervielfältigung des Programms ist nicht gestattet.<br>"
    "<span style='line-height:0;display:block;margin-top:17px;'>Erstellt von: Prof. Zebeli/Dr. Lucke</span>"
    "<span style='line-height:0;display:block;margin-top:17px;margin-bottom:0.3rem;'>Tierernährung Vetmeduni Wien</span>"
    "</div></div>",
    unsafe_allow_html=True,
)


# ============================================================
# CONSTANTS
# ============================================================
_VT_PASSWORD = "nutrition"

_NUTRIENTS_MAX3 = frozenset({"Vit. A", "Vit. D3", "Cu", "Zn", "Jod", "Selen"})

_DEFAULT_NUTRIENTS = frozenset({
    "Rp", "Rfe", "Ca", "P", "Mg", "K", "Na", "Cl", "Fe", "Cu", "Zn", "Mn",
    "Jod", "Selen", "Vit. A", "Vit. D3", "Vit. E", "Vit. B1", "Vit. B2",
    "Vit. B6", "Vit. B12", "Biotin", "Niacin", "Pantothensäure", "Folsäure",
    "Vit K", "Cholin",
})

# Patient selection hierarchy
_FALL_OPTIONS = {
    "Hund":  ["Erhaltung adult", "Trächtigkeit 2.Hälfte", "Laktation", "Aufzucht"],
    "Katze": ["Erhaltung", "Trächtigkeit", "Laktation", "Aufzucht"],
}

_DIAGNOSE_OPTIONS: dict = {
    ("Hund",  "Erhaltung adult"):       ["aktiv", "jung+aktiv", "Riesenrasse (aktiv)", "Terrier (aktiv)", "inaktiv", "senior+aktiv"],
    ("Hund",  "Trächtigkeit 2.Hälfte"): ["aktiv", "Riesenrasse (aktiv)", "Terrier (aktiv)", "inaktiv"],
    ("Hund",  "Laktation"):             ["1.Laktationswoche", "2.Laktationswoche", "3.Laktationswoche", "4.Laktationswoche"],
    ("Hund",  "Aufzucht"):              ["4.-14.Lebenswoche", ">14.Lebenswoche"],
    ("Katze", "Erhaltung"):             ["BCS≤5", "BCS>5"],
    ("Katze", "Trächtigkeit"):          None,   # no dropdown → hardcode "0"
    ("Katze", "Laktation"):             ["1.Laktationswoche", "2.Laktationswoche", "3.Laktationswoche",
                                         "4.Laktationswoche", "5.Laktationswoche", "6.Laktationswoche", "7.Laktationswoche"],
    ("Katze", "Aufzucht"):              None,   # no dropdown → hardcode "0"
}

# ============================================================
# PARSING: nutrient names from Bedarf (openpyxl, cached)
# ============================================================

@st.cache_data
def get_nutrient_names_from_bedarf(file_bytes: bytes, password: str = _VT_PASSWORD) -> list:
    """Read nutrient names from HilfstabelleMinMax row 1, cols F–BE (52 columns)."""
    buf = io.BytesIO(file_bytes)
    office = msoffcrypto.OfficeFile(buf)
    office.load_key(password=password)
    dec = io.BytesIO()
    office.decrypt(dec)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(dec, keep_vba=True, data_only=True)
    ws = wb["HilfstabelleMinMax"]
    names = []
    for c in range(6, 58):   # F=6 … BE=57  (52 columns)
        v = ws.cell(1, c).value
        raw = str(v).strip() if v is not None else f"col{c}"
        clean = re.sub(r"[_\s]*min\s*$", "", raw, flags=re.IGNORECASE).strip()
        names.append(clean if clean else raw)
    return names


# ============================================================
# READ PATIENT PARAMETERS FROM EXCEL (openpyxl, cached values)
# ============================================================

def _parse_identifier(identifier: str) -> tuple:
    """Parse 'Hund Erhaltung adult aktiv' → (animal, fall, diagnose).
    Uses known fall options to split correctly."""
    for animal in ["Hund", "Katze"]:
        if not identifier.startswith(animal + " "):
            continue
        rest = identifier[len(animal) + 1:]
        for fall in _FALL_OPTIONS.get(animal, []):
            if rest.startswith(fall + " ") or rest == fall:
                diagnose = rest[len(fall):].strip() or None
                return animal, fall, diagnose
        # fallback: split on first space
        parts = rest.split(" ", 1)
        return animal, parts[0], (parts[1].strip() if len(parts) > 1 else None)
    return None, None, None


@st.cache_data
def read_patient_params_from_excel(file_bytes: bytes, password: str = _VT_PASSWORD) -> dict:
    """Read patient parameters set by the student in the Berechnung sheet.

    H11 = computed identifier string (animal + life stage + condition).
    H12 = weight (kg).
    H10 = Lebenswoche (Aufzucht only).
    H14 = Welpen count (Laktation only).
    H15 = adult weight (Aufzucht only).
    Returns a dict with these values (None if not set / not applicable).
    """
    buf = io.BytesIO(file_bytes)
    office = msoffcrypto.OfficeFile(buf)
    office.load_key(password=password)
    dec = io.BytesIO()
    office.decrypt(dec)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(dec, keep_vba=True, data_only=True)
    ws = wb["Berechnung"]

    raw_id      = ws["H11"].value
    raw_weight  = ws["H12"].value
    raw_lw      = ws["H10"].value
    raw_welpen  = ws["H14"].value
    raw_adw     = ws["H15"].value

    def _float(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def _int(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    return {
        "identifier":   str(raw_id).strip() if raw_id is not None else "",
        "weight":       _float(raw_weight),
        "lebenswoche":  _int(raw_lw),
        "welpen":       _int(raw_welpen),
        "adult_weight": _float(raw_adw),
    }


def get_min_values_via_openpyxl(
    file_bytes: bytes,
    password: str,
    identifier: str,
    nutrient_names: list,
) -> tuple:
    """Read requirements from the Bedarf sheet using openpyxl (no Excel needed).
    Returns cached values from the last time the file was saved — weight cannot be
    recomputed here, so the returned values reflect whatever weight was active when
    the file was last saved in Excel."""
    buf = io.BytesIO(file_bytes)
    office = msoffcrypto.OfficeFile(buf)
    office.load_key(password=password)
    dec = io.BytesIO()
    office.decrypt(dec)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(dec, keep_vba=True, data_only=True)

    ws = wb["Bedarf"]

    target_row = None
    for r in range(2, 41):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip() == identifier.strip():
            target_row = r
            break
    if target_row is None:
        available = [str(ws.cell(row=r, column=1).value).strip()
                     for r in range(2, 41) if ws.cell(row=r, column=1).value]
        raise ValueError(f"Identifier '{identifier}' not found in Bedarf sheet. "
                         f"Available: {available}")

    min_col: dict = {}
    max_col: dict = {}
    # Also build a map of abbreviated name → full Bedarf name for cols with no suffix
    # (e.g. HilfstabelleMinMax uses "EPA + DHA" but Bedarf _min uses the full German name).
    no_suffix: dict = {}   # abbreviated col-64–117 name → its 1-indexed column number
    for c in range(1, 221):
        hdr = ws.cell(row=1, column=c).value
        if hdr is None:
            continue
        s = str(hdr).strip()
        clean_min = re.sub(r"[_\s]*min\s*$", "", s, flags=re.IGNORECASE).strip()
        if clean_min and clean_min != s:
            min_col[clean_min] = c
            continue
        clean_max = re.sub(r"[_\s]*max\s*$", "", s, flags=re.IGNORECASE).strip()
        if clean_max and clean_max != s:
            max_col[clean_max] = c
            continue
        # No suffix — store as potential alias (cols 64-117 use abbreviated names)
        no_suffix[s] = c

    # Build alias: abbreviated name (from HilfstabelleMinMax) → full Bedarf _min name.
    # Needed for nutrients like "EPA + DHA" whose full name differs between the two sheets.
    # Strategy: for each no-suffix column, find the corresponding _min column by position
    # (they are offset by exactly 52 columns: col N has _min at col N-52 and _max at N+52).
    _NO_SUFFIX_OFFSET = 52   # cols 64-117 are 52 cols after the _min block (cols 12-63)
    abbrev_to_full: dict = {}
    for abbr, c_abbr in no_suffix.items():
        c_min = c_abbr - _NO_SUFFIX_OFFSET
        if c_min >= 12:
            hdr_min = ws.cell(row=1, column=c_min).value
            if hdr_min:
                full = re.sub(r"[_\s]*min\s*$", "", str(hdr_min).strip(), flags=re.IGNORECASE).strip()
                if full != abbr and full in min_col:
                    abbrev_to_full[abbr] = full

    def _val(col_map, name):
        # Try direct match first, then alias lookup
        c = col_map.get(name) or col_map.get(abbrev_to_full.get(name, ""))
        if c is None:
            return 0.0
        v = ws.cell(row=target_row, column=c).value
        try:
            fv = float(v)
            return 0.0 if math.isnan(fv) else fv
        except (TypeError, ValueError):
            return 0.0

    mins = {name: _val(min_col, name) for name in nutrient_names}
    maxs = {name: _val(max_col, name) for name in nutrient_names if _val(max_col, name) > 0}
    return mins, maxs


def build_constraints_from_mins(
    mins_dict: dict,
    selected_nutrients: list,
    maxs_dict: Optional[dict] = None,
    use_calculated_max: bool = False,
) -> pd.DataFrame:
    """Build constraints DataFrame from Excel-calculated min values.
    use_calculated_max=True → use maxs_dict (HilfstabelleMinMax row 16).
    use_calculated_max=False → Estimated: min × 3 or × 5.
    """
    constraints = pd.DataFrame(
        index=["Tagesbedarf", "Maximaler_Wert", "Grundnahrung"],
        columns=selected_nutrients,
        dtype=float,
    )
    for n in selected_nutrients:
        mn = float(mins_dict.get(n, 0.0) or 0.0)
        if use_calculated_max and maxs_dict:
            raw = maxs_dict.get(n, None)
            try:
                mx = float(raw)
                if math.isnan(mx) or mx <= 0:
                    mx = np.nan
            except (TypeError, ValueError):
                mx = np.nan
        else:
            mx = mn * (3.0 if n in _NUTRIENTS_MAX3 else 5.0) if mn > 0 else np.nan
        constraints.loc["Tagesbedarf",   n] = mn
        constraints.loc["Maximaler_Wert", n] = mx
        constraints.loc["Grundnahrung",   n] = 0.0
    return constraints


# ============================================================
# PARSING: Futtermittel (base diet foods from volle_Tabelle)
# ============================================================

@st.cache_data
def parse_futtermittel_sheet(file_bytes: bytes, password: str = _VT_PASSWORD) -> pd.DataFrame:
    """Read Futtermittel sheet. Returns: Identifier | Typ | Kategorie | Name | <nutrient cols per 100g>."""
    buf = io.BytesIO(file_bytes)
    office = msoffcrypto.OfficeFile(buf)
    office.load_key(password=password)
    dec = io.BytesIO()
    office.decrypt(dec)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(dec, keep_vba=True, data_only=True)
    ws = wb["Futtermittel"]

    all_rows = list(ws.values)
    if len(all_rows) < 4:
        return pd.DataFrame()

    # Row 3 (0-indexed 2): headers; nutrient cols M–CT = 0-indexed 12–97
    NUTR_S, NUTR_E = 12, 97
    hrow = all_rows[2]

    def _clean_ft(h: str) -> str:
        return re.sub(r"\s*\[.*", "", str(h)).strip()

    # Build one cleaned header per column position (no dedup — keep all).
    # When two columns share the same cleaned name (e.g. "Vit. A [μg]" and
    # "Vit. A [IE]" both become "Vit. A"), the Berechnung sheet always uses
    # the LAST occurrence, so we apply a "last wins" write in the data loop.
    clean_hdrs = []
    for i in range(NUTR_S, NUTR_E + 1):
        raw = str(hrow[i]).strip() if (i < len(hrow) and hrow[i] is not None) else f"col{i+1}"
        clean_hdrs.append(_clean_ft(raw))

    # Ca, P, K are stored as mg/100g in Futtermittel but the Bedarf sheet
    # uses grams for these three — divide by 1000 so units align.
    _MG_TO_G = {"Ca", "P", "K"}

    rows = []
    for tpl in all_rows[3:]:
        if len(tpl) <= 11:
            continue
        identifier = tpl[3]    # col D
        name       = tpl[11]   # col L
        if identifier is None or name is None:
            continue
        typ       = tpl[4] if len(tpl) > 4 else None   # col E
        spezies   = tpl[5] if len(tpl) > 5 else None   # col F
        kategorie = tpl[6] if len(tpl) > 6 else None   # col G
        num       = tpl[8] if len(tpl) > 8 else None   # col I = Nummer fortlaufend

        _preis_raw = tpl[98] if len(tpl) > 98 else None  # col CU = Preis (€) pro kg
        try:
            _preis = float(_preis_raw) if _preis_raw is not None else None
        except (TypeError, ValueError):
            _preis = None

        rdata: dict = {
            "Identifier":  str(identifier).strip(),
            "Typ":         str(typ).strip()       if typ       else "",
            "Spezies":     str(spezies).strip()   if spezies   else "",
            "Kategorie":   str(kategorie).strip() if kategorie else "",
            "Num":         int(num) if num is not None else -1,
            "Name":        str(name).strip(),
            "Preis (€/kg)": _preis,
        }
        for j, ch in enumerate(clean_hdrs):
            idx = NUTR_S + j
            v   = tpl[idx] if idx < len(tpl) else None
            try:
                raw_val = float(v) if v is not None else 0.0
                rdata[ch] = raw_val * 0.001 if ch in _MG_TO_G else raw_val
            except (TypeError, ValueError):
                rdata[ch] = 0.0
        rows.append(rdata)

    return pd.DataFrame(rows) if rows else pd.DataFrame()


@st.cache_data
def get_futtermittel_nutrient_units(file_bytes: bytes, password: str = _VT_PASSWORD) -> dict:
    """Return {nutrient_name: display_unit} from Futtermittel column headers (last-wins for duplicates).
    Ca/P/K are stored in g (after /1000 conversion) so their unit is reported as 'g'."""
    buf = io.BytesIO(file_bytes)
    office = msoffcrypto.OfficeFile(buf)
    office.load_key(password=password)
    dec = io.BytesIO()
    office.decrypt(dec)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(dec, keep_vba=True, data_only=True)
    hrow = list(wb["Futtermittel"].values)[2]

    _MG_TO_G = {"Ca", "P", "K"}
    units: dict = {}
    for i in range(12, 98):
        raw = str(hrow[i]).strip() if (i < len(hrow) and hrow[i] is not None) else ""
        if not raw:
            continue
        name = re.sub(r"\s*\[.*", "", raw).strip()
        m    = re.search(r"\[([^\]/]+)", raw)
        unit = m.group(1).strip() if m else ""
        if name:
            units[name] = "g" if name in _MG_TO_G else unit
            short_name = re.sub(r'\s*\([^)]*\)', '', name).strip()
            if short_name and short_name != name:
                units.setdefault(short_name, "g" if name in _MG_TO_G else unit)
    return units


@st.cache_data
def get_futtermittel_column_names(file_bytes: bytes, password: str = _VT_PASSWORD) -> list:
    """Return deduplicated list of cleaned Futtermittel column names (cols 12-97, in order)."""
    buf = io.BytesIO(file_bytes)
    office = msoffcrypto.OfficeFile(buf)
    office.load_key(password=password)
    dec = io.BytesIO()
    office.decrypt(dec)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(dec, keep_vba=True, data_only=True)
    hrow = list(wb["Futtermittel"].values)[2]
    seen, names = set(), []
    for i in range(12, 98):
        raw  = str(hrow[i]).strip() if (i < len(hrow) and hrow[i] is not None) else ""
        name = re.sub(r"\s*\[.*", "", raw).strip() if raw else ""
        if name and name not in seen:
            names.append(name)
            seen.add(name)
    return names


# ============================================================
# PARSING: Supplements
# ============================================================

@st.cache_data
def parse_supplements_excel(file) -> pd.DataFrame:
    df_efm   = pd.read_excel(file, sheet_name="EFM",                header=2)
    df_einzel = pd.read_excel(file, sheet_name="Einzelfuttermittel", header=2)
    df_einzel = df_einzel.rename(columns={"Taurin [mg]/[100 g]": "Taurin [mg]/[100g]"})
    df_efm    = df_efm.dropna(axis=1, how="all").dropna(subset=["Identifier"])
    df_einzel = df_einzel.dropna(axis=1, how="all").dropna(subset=["Identifier"])
    df_efm_slim   = df_efm.iloc[:, 4:-13]
    df_einzel_slim = df_einzel.iloc[:, 5:-12]
    for c in set(df_efm_slim.columns) - set(df_einzel_slim.columns):
        df_einzel_slim[c] = 0
    for c in set(df_einzel_slim.columns) - set(df_efm_slim.columns):
        df_efm_slim[c] = 0
    df = pd.concat([df_efm_slim, df_einzel_slim], ignore_index=True)
    df.columns = [str(c).strip().replace("100 g", "100g") for c in df.columns]
    return df


def validate_supplements_df(supplements: pd.DataFrame):
    issues = []
    required_cols = {"Futtermittel", "Preis (€) pro kg"}
    missing = required_cols - set(supplements.columns)
    if missing:
        issues.append(f"Required columns missing: {', '.join(sorted(missing))}.")
    if supplements.shape[0] == 0:
        issues.append("No rows detected after parsing.")
    if "Preis (€) pro kg" in supplements.columns:
        prices = pd.to_numeric(supplements["Preis (€) pro kg"], errors="coerce")
        if prices.isna().any():
            issues.append("Non-numeric or missing prices detected.")
        if (prices < 0).any():
            issues.append("Negative prices detected.")
    return issues


def get_excluded_supplements(supplements: pd.DataFrame) -> pd.DataFrame:
    if "Preis (€) pro kg" not in supplements.columns or "Futtermittel" not in supplements.columns:
        return pd.DataFrame(columns=["Excel row", "Supplement", "Price (€) per kg", "Exclusion reason"])
    df = supplements.copy()
    df["_row"] = df.index + 2
    prices = pd.to_numeric(df["Preis (€) pro kg"], errors="coerce")
    excl = df.loc[prices.isna(), ["_row", "Futtermittel", "Preis (€) pro kg"]].copy()
    excl["reason"] = "No valid price provided"
    excl = excl.rename(columns={"_row": "Excel row", "Futtermittel": "Supplement",
                                  "Preis (€) pro kg": "Price (€) per kg", "reason": "Exclusion reason"})
    return excl.sort_values(["Excel row", "Supplement"]).reset_index(drop=True)


# ============================================================
# EDITABLE TABLE HELPERS
# ============================================================

def constraints_to_edit_df(constraints: pd.DataFrame) -> pd.DataFrame:
    nutrients = list(constraints.columns)
    edit = pd.DataFrame({
        "Nährstoff":        nutrients,
        "Tagesbedarf (Min)": pd.to_numeric(constraints.loc["Tagesbedarf",    nutrients], errors="coerce"),
        "Maximalwert (Max)": pd.to_numeric(constraints.loc["Maximaler_Wert", nutrients], errors="coerce"),
        "Grundnahrung":      pd.to_numeric(constraints.loc["Grundnahrung",   nutrients], errors="coerce"),
    })
    edit["Bedarf nach Grundnahrung (Min-Base)"] = (edit["Tagesbedarf (Min)"] - edit["Grundnahrung"]).clip(lower=0.0)
    edit["🗑 Löschen"] = False
    return edit


def edit_df_to_constraints(edit_df: pd.DataFrame, original_constraints: pd.DataFrame) -> pd.DataFrame:
    nutrients = edit_df["Nährstoff"].tolist()
    out = original_constraints.copy()
    for n in nutrients:
        if n not in out.columns:
            out[n] = np.nan
    out = out.reindex(columns=nutrients)
    out.loc["Tagesbedarf"]    = pd.to_numeric(edit_df["Tagesbedarf (Min)"].values, errors="coerce")
    out.loc["Maximaler_Wert"] = pd.to_numeric(edit_df["Maximalwert (Max)"].values, errors="coerce")
    out.loc["Grundnahrung"]   = pd.to_numeric(edit_df["Grundnahrung"].values,       errors="coerce")
    return out


# ============================================================
# SUPPLEMENT BUILD + OPTIMIZATION
# ============================================================

def canonical_name(col: str) -> str:
    # Only split at "/" that precedes a unit denominator (kg, 100g) — NOT at "/"
    # that is part of the nutrient name itself (e.g. "Ca/P Verhältnis").
    m = re.split(r"\s*/\s*\[?(100\s*g|kg)\]?(?:\s|$)", str(col), maxsplit=1, flags=re.IGNORECASE)
    return m[0].strip()

def is_100g_col(col: str) -> bool:
    return re.search(r"100\s*g", str(col), flags=re.IGNORECASE) is not None

def infer_available_nutrients_from_supplements(supplements: pd.DataFrame) -> set:
    ignore = {"Futtermittel", "Preis (€) pro kg", "Identifier"}
    return {canonical_name(c) for c in supplements.columns if c not in ignore}

def build_supp_clean(supplements: pd.DataFrame, nutrient_cols: list) -> pd.DataFrame:
    df = supplements.copy()
    df = df.dropna(subset=["Futtermittel", "Preis (€) pro kg"]).copy()
    df["Preis (€) pro kg"] = pd.to_numeric(df["Preis (€) pro kg"], errors="coerce")
    df = df.dropna(subset=["Preis (€) pro kg"])
    all_cols = [c for c in df.columns if c not in ["Futtermittel", "Preis (€) pro kg"]]
    canon_map: dict = {}
    for c in all_cols:
        canon_map.setdefault(canonical_name(c), []).append(c)
    out = pd.DataFrame({"Futtermittel": df["Futtermittel"], "Preis (€) pro kg": df["Preis (€) pro kg"]})
    for n in nutrient_cols:
        cols_for_n = canon_map.get(n, [])
        if not cols_for_n:
            out[n] = 0.0
            continue
        vals = []
        for c in cols_for_n:
            s = pd.to_numeric(df[c], errors="coerce").replace([np.inf, -np.inf], np.nan).fillna(0.0)
            if is_100g_col(c):
                s = s * 10.0
            vals.append(s)
        out[n] = pd.concat(vals, axis=1).mean(axis=1)
    agg = {n: "mean" for n in nutrient_cols}
    agg["Preis (€) pro kg"] = "min"
    out = out.groupby("Futtermittel", as_index=True).agg(agg)
    out = out[out[nutrient_cols].abs().sum(axis=1) > 0]
    return out


def optimize_fast(constraints_effective: pd.DataFrame, supplements: pd.DataFrame,
                  max_supplements: Optional[int] = None,
                  excluded_supplements: Optional[list] = None):
    if excluded_supplements:
        supplements = supplements[~supplements["Futtermittel"].isin(set(excluded_supplements))].copy()
    nutrients_from_constraints = [c for c in constraints_effective.columns if "Verhältnis" not in str(c)]
    available = infer_available_nutrients_from_supplements(supplements)
    base_all = pd.to_numeric(constraints_effective.loc["Grundnahrung",    nutrients_from_constraints], errors="coerce").fillna(0.0)
    min_all  = pd.to_numeric(constraints_effective.loc["Tagesbedarf",     nutrients_from_constraints], errors="coerce").fillna(0.0)
    max_all  = pd.to_numeric(constraints_effective.loc["Maximaler_Wert",  nutrients_from_constraints], errors="coerce")
    min_supp_all = (min_all - base_all).clip(lower=0.0)
    missing_required = [n for n in nutrients_from_constraints
                        if (min_supp_all.get(n, 0.0) > 0) and (n not in available)]
    if missing_required:
        return "Infeasible", None, None, {
            "Missing nutrients in supplement database": missing_required,
            "Note": "These nutrients have a remaining requirement (>0) but no matching column in the supplement DB.",
        }, {"reason": "missing_required_cols"}
    nutrient_cols = [n for n in nutrients_from_constraints if n in available]
    base     = base_all.reindex(nutrient_cols).fillna(0.0)
    min_req  = min_all.reindex(nutrient_cols).fillna(0.0)
    max_req  = max_all.reindex(nutrient_cols)
    min_supp = (min_req - base).clip(lower=0.0)
    max_supp = max_req - base
    infeasible_base_over_max = max_supp[max_supp < 0]
    infeasible_msg = infeasible_base_over_max.to_dict() if len(infeasible_base_over_max) else None
    supp_clean = build_supp_clean(supplements, nutrient_cols)
    if supp_clean.shape[0] == 0:
        return "Infeasible", None, None, {"Note": "No supplements remain after cleaning."}, {"reason": "empty_supp_clean"}
    MIN_KG, BIG_M_KG = 1.0 / 1000.0, 10.0

    def _diagnose():
        diag_rows = []
        for n in nutrient_cols:
            coeff = supp_clean[n].astype(float)
            pos_sum = float(np.maximum(coeff.values, 0.0).sum())
            max_possible = BIG_M_KG * pos_sum
            min_needed = float(min_supp.get(n, 0.0))
            max_allowed_val = max_supp.get(n, np.nan)
            max_allowed = float(max_allowed_val) if pd.notna(max_allowed_val) else np.nan
            has_positive_source = bool((coeff > 0).any())
            top_sources = coeff.sort_values(ascending=False).head(8)
            top_sources_dict = {k: float(v) for k, v in top_sources.items() if float(v) != 0.0}
            one_g_effect_dict = {k: float(v) * MIN_KG for k, v in top_sources_dict.items()}
            reason_flags = []
            if min_needed > 0 and not has_positive_source:
                reason_flags.append("Min > 0 but no positive source available")
            if min_needed > max_possible + 1e-12:
                reason_flags.append("Min exceeds max achievable value (BIG_M limit)")
            if pd.notna(max_allowed) and max_allowed < 0:
                reason_flags.append("Base-diet exceeds Max (Max-Base < 0)")
            diag_rows.append({
                "Nutrient": n, "Min_After_Base_Diet": min_needed,
                "Max_After_Base_Diet": max_allowed if pd.notna(max_allowed) else np.nan,
                "Max_Achievable_Rough": max_possible, "Has_Positive_Source": has_positive_source,
                "Flags": " | ".join(reason_flags),
                "Top_Sources_coeff_per_kg": top_sources_dict,
                "1g_Effect_TopSources": one_g_effect_dict,
            })
        diag_df = pd.DataFrame(diag_rows)
        likely = diag_df[(diag_df["Flags"].astype(str).str.len() > 0) |
                         (diag_df["Min_After_Base_Diet"] > diag_df["Max_Achievable_Rough"] + 1e-12)].copy()
        return {
            "diag_df": diag_df,
            "likely_problem_nutrients": likely.sort_values("Min_After_Base_Diet", ascending=False).reset_index(drop=True),
            "MIN_KG": MIN_KG, "BIG_M_KG": BIG_M_KG,
            "nutrient_cols": nutrient_cols, "supp_clean_shape": tuple(supp_clean.shape),
        }

    if infeasible_msg:
        return "Infeasible", None, None, {"Base-Diet over Max": infeasible_msg}, {
            "nutrient_cols": nutrient_cols, "supp_clean": supp_clean,
            "base": base, "min_req": min_req, "max_req": max_req, "diagnose": _diagnose(),
        }
    model = pulp.LpProblem("Supplement_Opt", pulp.LpMinimize)
    x = {s: pulp.LpVariable(f"x_{i}", lowBound=0) for i, s in enumerate(supp_clean.index)}
    y = {s: pulp.LpVariable(f"y_{i}", cat="Binary") for i, s in enumerate(supp_clean.index)}
    model += pulp.lpSum(supp_clean.loc[s, "Preis (€) pro kg"] * x[s] for s in supp_clean.index)
    if max_supplements is not None:
        model += pulp.lpSum(y[s] for s in supp_clean.index) <= max_supplements
    for s in supp_clean.index:
        model += x[s] <= BIG_M_KG * y[s]
        model += x[s] >= MIN_KG * y[s]
    for n in nutrient_cols:
        intake = pulp.lpSum(supp_clean.loc[s, n] * x[s] for s in supp_clean.index)
        if float(min_supp.get(n, 0)) > 0:
            model += intake >= float(min_supp[n])
        if pd.notna(max_supp.get(n, np.nan)):
            model += intake <= float(max_supp[n])
    model.solve(pulp.PULP_CBC_CMD(msg=False))
    status = pulp.LpStatus[model.status]
    dbg = {"nutrient_cols": nutrient_cols, "supp_clean": supp_clean,
           "base": base, "min_req": min_req, "max_req": max_req, "diagnose": _diagnose()}
    if status != "Optimal":
        dbg["solver_status"] = status
        return status, None, None, infeasible_msg, dbg
    solution = pd.Series({s: x[s].value() for s in x if (x[s].value() or 0) > 1e-9}).sort_values(ascending=False)
    cost = float(pulp.value(model.objective))
    return status, solution, cost, infeasible_msg, dbg


def optimize_with_slacks(constraints_effective: pd.DataFrame, supplements: pd.DataFrame,
                         penalty=1e6, excluded_supplements: Optional[list] = None):
    if excluded_supplements:
        supplements = supplements[~supplements["Futtermittel"].isin(set(excluded_supplements))].copy()
    nutrients = [c for c in constraints_effective.columns if "Verhältnis" not in str(c)]
    available = infer_available_nutrients_from_supplements(supplements)
    base     = pd.to_numeric(constraints_effective.loc["Grundnahrung",   nutrients], errors="coerce").fillna(0.0)
    minv     = pd.to_numeric(constraints_effective.loc["Tagesbedarf",    nutrients], errors="coerce").fillna(0.0)
    maxv     = pd.to_numeric(constraints_effective.loc["Maximaler_Wert", nutrients], errors="coerce")
    min_supp = (minv - base).clip(lower=0.0)
    max_supp = maxv - base
    nutrient_cols = [n for n in nutrients if n in available]
    supp_clean = build_supp_clean(supplements, nutrient_cols)
    MIN_KG, BIG_M_KG = 1.0 / 1000.0, 10.0
    model = pulp.LpProblem("Debug_Slack", pulp.LpMinimize)
    x     = {s: pulp.LpVariable(f"x_{i}", lowBound=0) for i, s in enumerate(supp_clean.index)}
    y     = {s: pulp.LpVariable(f"y_{i}", cat="Binary") for i, s in enumerate(supp_clean.index)}
    s_min = {n: pulp.LpVariable(f"smin_{i}", lowBound=0) for i, n in enumerate(nutrient_cols)}
    s_max = {n: pulp.LpVariable(f"smax_{i}", lowBound=0) for i, n in enumerate(nutrient_cols)}
    model += (penalty * pulp.lpSum(s_min[n] + s_max[n] for n in nutrient_cols)
              + pulp.lpSum(supp_clean.loc[s, "Preis (€) pro kg"] * x[s] for s in supp_clean.index))
    for s in supp_clean.index:
        model += x[s] <= BIG_M_KG * y[s]
        model += x[s] >= MIN_KG * y[s]
    for n in nutrient_cols:
        intake = pulp.lpSum(supp_clean.loc[s, n] * x[s] for s in supp_clean.index)
        model += intake + s_min[n] >= float(min_supp.get(n, 0.0))
        if pd.notna(max_supp.get(n)):
            model += intake - s_max[n] <= float(max_supp[n])
    model.solve(pulp.PULP_CBC_CMD(msg=False))
    rows = [{"Nutrient": n, "Slack_Min": float(s_min[n].value() or 0), "Slack_Max": float(s_max[n].value() or 0)}
            for n in nutrient_cols]
    return pd.DataFrame(rows).sort_values(["Slack_Min", "Slack_Max"], ascending=False)


# ============================================================
# MATCHING HELPERS
# ============================================================

def normalize_name(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^\w\s\[\]\(\)\-\/\+\.]", "", s)
    return s.strip()

def strip_unit_brackets(s: str) -> str:
    """Remove [unit] suffixes from a nutrient name, e.g. 'Calcium [g]' → 'Calcium'."""
    return re.sub(r"\s*\[.*?\]", "", str(s)).strip()

def extract_supp_unit(canonical: str) -> str:
    """Extract unit string from a canonical supplement name, e.g. 'Calcium [g]' → 'g'."""
    m = re.search(r"\[([^\]]+)\]", str(canonical))
    return m.group(1).strip() if m else ""

_MASS_BASE_G = {
    "kg": 1e3, "g": 1.0, "mg": 1e-3,
    "µg": 1e-6, "μg": 1e-6, "ug": 1e-6, "mcg": 1e-6,
    "ng": 1e-9,
}
_ENERGY_BASE_KCAL = {
    "kcal": 1.0, "cal": 1e-3, "kj": 1 / 4.184, "mj": 1000 / 4.184,
}

def get_unit_conversion_factor(req_unit: str, supp_unit: str):
    """Return factor to multiply requirement values by so they match supp_unit scale.
    Returns 1.0 if units are the same or either is missing.
    Returns a float if convertible.
    Returns None if the units are incompatible (different dimensions)."""
    ru = (req_unit or "").strip().lower()
    su = (supp_unit or "").strip().lower()
    if not ru or not su or ru == su:
        return 1.0
    if ru in _MASS_BASE_G and su in _MASS_BASE_G:
        return _MASS_BASE_G[ru] / _MASS_BASE_G[su]
    if ru in _ENERGY_BASE_KCAL and su in _ENERGY_BASE_KCAL:
        return _ENERGY_BASE_KCAL[ru] / _ENERGY_BASE_KCAL[su]
    return None

def build_initial_mapping_and_status(ration_nutrients: list, supp_names: list, req_units: dict = None):
    # Full-name lookup (includes unit suffix, e.g. "taurin [mg]")
    supp_norm_full = {normalize_name(s): s for s in supp_names}

    # Base-name → list of canonicals (multiple units possible per nutrient name)
    supp_base_to_canonicals: dict = {}
    for s in supp_names:
        base = normalize_name(strip_unit_brackets(s))
        supp_base_to_canonicals.setdefault(base, []).append(s)

    def _pick_best(candidates: list, req_unit: str) -> str:
        """From multiple supplement canonicals for the same base name, prefer unit match."""
        if len(candidates) == 1:
            return candidates[0]
        if req_unit:
            ru = req_unit.lower()
            for c in candidates:
                if extract_supp_unit(c).lower() == ru:
                    return c
        return candidates[0]

    mapping, status, mode = {}, {}, {}
    for r in ration_nutrients:
        rn       = normalize_name(r)
        rn_base  = normalize_name(strip_unit_brackets(r))
        req_unit = (req_units or {}).get(r, "")

        if rn in supp_norm_full:
            mapping[r] = supp_norm_full[rn]; status[r] = "exact"; mode[r] = "auto"
        elif rn_base in supp_base_to_canonicals:
            mapping[r] = _pick_best(supp_base_to_canonicals[rn_base], req_unit)
            status[r] = "exact"; mode[r] = "auto"
        else:
            close_base = difflib.get_close_matches(rn_base, list(supp_base_to_canonicals.keys()), n=1, cutoff=0.78)
            close_full = difflib.get_close_matches(rn,      list(supp_norm_full.keys()),           n=1, cutoff=0.78)
            if close_base:
                mapping[r] = _pick_best(supp_base_to_canonicals[close_base[0]], req_unit)
                status[r] = "fuzzy"; mode[r] = "auto"
            elif close_full:
                mapping[r] = supp_norm_full[close_full[0]]; status[r] = "fuzzy"; mode[r] = "auto"
            else:
                mapping[r] = None; status[r] = "missing"; mode[r] = "auto"
    return mapping, status, mode

def apply_mapping_to_constraints(
    constraints_effective: pd.DataFrame,
    mapping: dict,
    req_units: dict = None,
) -> pd.DataFrame:
    cols = list(constraints_effective.columns)
    rename_map, keep_cols, scale_map = {}, [], {}
    for c in cols:
        if c in mapping:
            target = mapping.get(c)
            if target is None:
                continue
            rename_map[c] = target
            if req_units:
                factor = get_unit_conversion_factor(
                    req_units.get(c, ""), extract_supp_unit(target)
                )
                if factor is not None and factor != 1.0:
                    scale_map[c] = factor
        keep_cols.append(c)
    out = constraints_effective[keep_cols].copy()
    for c, factor in scale_map.items():
        out[c] = pd.to_numeric(out[c], errors="coerce") * factor
    out = out.rename(columns=rename_map)
    if out.columns.duplicated().any():
        out = out.T.groupby(level=0).max().T
    return out


def build_futtermittel_nutrient_mapping(req_names: list, ft_col_names: list) -> tuple:
    """Map requirement names (short, from HilfstabelleMinMax) to Futtermittel column names.
    Returns (mapping, status):
      mapping: {req_name → ft_col_name or None}
      status:  {req_name → "exact"|"base"|"fuzzy"|"missing"}
    "base" means matched after stripping parentheticals, e.g. "Vit. E" → "Vit. E (a-Tocopherol)".
    """
    def _strip_paren(s: str) -> str:
        return re.sub(r'\s*\([^)]*\)', '', str(s)).strip()

    ft_norm_full: dict = {normalize_name(c): c for c in ft_col_names}
    ft_base_to_cols: dict = {}
    for c in ft_col_names:
        base = normalize_name(_strip_paren(c))
        ft_base_to_cols.setdefault(base, []).append(c)

    mapping, status = {}, {}
    for r in req_names:
        rn      = normalize_name(r)
        rn_base = normalize_name(_strip_paren(r))
        if rn in ft_norm_full:
            mapping[r] = ft_norm_full[rn]; status[r] = "exact"
        elif rn_base in ft_base_to_cols:
            mapping[r] = ft_base_to_cols[rn_base][-1]; status[r] = "base"
        else:
            close = difflib.get_close_matches(rn, list(ft_norm_full.keys()), n=1, cutoff=0.65)
            if close:
                mapping[r] = ft_norm_full[close[0]]; status[r] = "fuzzy"
            else:
                mapping[r] = None; status[r] = "missing"
    return mapping, status


def _lookup_base_total(base_totals: dict, nutrient_name: str, ft_mapping: dict = None) -> float:
    """Return base diet total for nutrient_name: mapping → exact → normalized → fuzzy fallback."""
    if ft_mapping:
        ft_col = ft_mapping.get(nutrient_name)
        if ft_col is not None and ft_col in base_totals:
            return float(base_totals[ft_col])
    v = base_totals.get(nutrient_name)
    if v is not None:
        return float(v)
    nn = normalize_name(nutrient_name)
    for k, kv in base_totals.items():
        if normalize_name(k) == nn:
            return float(kv)
    close = difflib.get_close_matches(nn, [normalize_name(k) for k in base_totals], n=1, cutoff=0.70)
    if close:
        for k, kv in base_totals.items():
            if normalize_name(k) == close[0]:
                return float(kv)
    return 0.0


# ============================================================
# 2) UPLOAD & SELECTION SECTION
# ============================================================
st.markdown("<div style='margin-top:1.2rem;'></div>", unsafe_allow_html=True)
st.markdown("### 📥 Upload Data & Select Nutrition Needs")

constraints_raw      = None
constraints_effective = None
supplements          = None
volle_tabelle_ok     = False
supp_ok              = False
excluded_supps_table = pd.DataFrame()

with st.expander("📥 Upload Data & Select Nutrition Needs (expand)", expanded=True):
    left, right = st.columns(2)

    # --------------------------------------------------------
    # LEFT: volle_Tabelle upload + patient inputs
    # --------------------------------------------------------
    with left:
        with st.expander("Nutrition Requirements", expanded=True):
            st.markdown("<div class='upload-title' style='margin-bottom:0.6rem;'>CarniDiet© File</div>", unsafe_allow_html=True)

            vt_file = st.file_uploader(
                "volle_Tabelle Upload", type=["xlsm", "xlsx"],
                label_visibility="collapsed", key="vt_uploader",
            )

            if vt_file:
                try:
                    file_bytes = vt_file.read()
                    st.session_state["v3_file_bytes"] = file_bytes

                    # Reset CarniDiet-specific state when a new file is uploaded.
                    # Supplement database and base diet state are intentionally preserved.
                    _new_hash = hashlib.md5(file_bytes).hexdigest()[:12]
                    if st.session_state.get("v3_last_file_hash") != _new_hash:
                        _carnidiet_keys = [
                            "v3_ft_map", "v3_ft_map_sig", "v3_ft_map_status", "v3_ft_map_locked",
                            "v3_constraints_raw", "v3_calc_sig", "v3_mins_dict", "v3_maxs_dict",
                            "v3_nutrient_selection", "v3_max_method",
                            "constraints_edit_df", "constraints_locked",
                            "constraints_effective_df", "constraints_editor_nonce",
                            "nutrient_mapping", "nutrient_mapping_status",
                            "nutrient_mapping_mode", "nutrient_mapping_signature", "mapping_locked",
                        ]
                        for _k in _carnidiet_keys:
                            st.session_state.pop(_k, None)
                        st.session_state["v3_last_file_hash"] = _new_hash
                        st.rerun()

                    nutrient_names = get_nutrient_names_from_bedarf(file_bytes)
                    st.markdown("<div class='okrow'>✅ File loaded</div>", unsafe_allow_html=True)

                    # ── Info: parameters come from Excel ─────────────
                    st.markdown(
                        "<div style='margin-top:0.7rem;background:rgba(255,140,0,0.10);border:1.5px solid rgba(255,140,0,0.70);"
                        "border-radius:0.5rem;padding:0.65rem 1rem;font-size:0.95rem;font-weight:600;"
                        "color:#7a4400;margin-bottom:0.4rem;'>"
                        "🔶 Key variables such as Spezies, Bedarf, Gewicht and their related inputs must be set directly "
                        "in the CarniDiet© File before uploading. "
                        "To change these settings, update them in Excel and re-upload the file."
                        "</div>",
                        unsafe_allow_html=True,
                    )

                    # ── Read parameters from the uploaded file ────────
                    params = read_patient_params_from_excel(file_bytes)
                    identifier = params["identifier"]

                    if not identifier:
                        st.error(
                            "❌ Could not read patient parameters from the file. "
                            "Make sure you have set the parameters in Excel and saved the file before uploading."
                        )
                        selected_nutrients = []
                    else:
                        animal, selected_fall, selected_diagnose = _parse_identifier(identifier)
                        is_aufzucht  = (selected_fall == "Aufzucht")
                        is_laktation = (selected_fall == "Laktation")

                        # ── Read-only parameter display ───────────────
                        st.markdown("#### 📋 Parameters (from file)")
                        _param_rows = [
                            ("🐾 Animal",      animal or "—"),
                            ("📋 Life Stage",  selected_fall or "—"),
                        ]
                        if selected_diagnose and selected_diagnose.strip() not in ("0", ""):
                            _param_rows.append(("🔎 Condition", selected_diagnose))

                        _wlabel = "🐶 Puppy Weight" if is_aufzucht else "⚖️ Weight"
                        _wval   = f"{params['weight']:.2f} kg" if params["weight"] is not None else "—"
                        _param_rows.append((_wlabel, _wval))

                        if is_aufzucht and params["lebenswoche"] is not None:
                            _param_rows.append(("📅 Week of Life", str(params["lebenswoche"])))
                        if is_aufzucht and params["adult_weight"] is not None:
                            _param_rows.append(("⚖️ Adult Weight", f"{params['adult_weight']:.1f} kg"))
                        if is_laktation and params["welpen"] is not None:
                            _param_rows.append(("🐕 Puppies", str(params["welpen"])))

                        _param_html = (
                            "<div style='background:rgba(28,131,225,0.08);border:1px solid "
                            "rgba(28,131,225,0.25);border-radius:0.5rem;padding:0.6rem 1rem;"
                            "font-size:0.95rem;margin-bottom:0.5rem;'>"
                        )
                        for _lbl, _val in _param_rows:
                            _param_html += (
                                f"<div style='padding:0.18rem 0;'>"
                                f"<span style='font-weight:600;color:#19425e;'>{_lbl}:</span>"
                                f"&nbsp;{_val}</div>"
                            )
                        _param_html += "</div>"
                        st.markdown(_param_html, unsafe_allow_html=True)

                    # ── Nutrient selection ────────────────────────────
                    st.markdown("<div style='margin-top:1.1rem;'></div>", unsafe_allow_html=True)
                    st.markdown("#### 🧪 Which nutrients to include?")
                    st.markdown(
                        "<div style='margin-top:-0.4rem;margin-bottom:0.5rem;"
                        "font-size:0.9rem;opacity:0.75;'>All 52 nutrients from the Bedarf sheet are available.</div>",
                        unsafe_allow_html=True,
                    )

                    default_in_data = [n for n in nutrient_names if n in _DEFAULT_NUTRIENTS]
                    extra_in_data   = [n for n in nutrient_names if n not in _DEFAULT_NUTRIENTS]
                    # Pull Pantothensäure out — rendered in its own row below the main grid.
                    _panto = "Pantothensäure"
                    _default_main = [n for n in default_in_data if n != _panto]
                    _panto_present = _panto in default_in_data

                    if "v3_nutrient_selection" not in st.session_state:
                        st.session_state["v3_nutrient_selection"] = set(default_in_data)

                    sel = st.session_state["v3_nutrient_selection"]

                    # Normalise checkbox label font size across narrow (4-col) and
                    # wider (3-col) grid sections so they look identical.
                    st.markdown(
                        "<style>"
                        "div[data-testid='stCheckbox'] label p {"
                        "  font-size: 1rem !important;"
                        "}"
                        "div[data-testid='stExpander'] summary span p {"
                        "  font-size: 1.05rem !important;"
                        "}"
                        "</style>",
                        unsafe_allow_html=True,
                    )
                    with st.expander("Select nutrients (click to expand)", expanded=True):
                        st.markdown(
                            "<div style='font-size:1.05rem;font-weight:700;color:#19425e;"
                            "margin-bottom:0.4rem;'>📌 Standard nutrients (default)</div>",
                            unsafe_allow_html=True,
                        )
                        # Render main grid 4 per row (always 4 cols so rows stay aligned).
                        # Detect whether Pantothensäure can share the last row (needs 2 free slots).
                        _chunks = [_default_main[i:i+4] for i in range(0, len(_default_main), 4)]
                        _merge_panto = _panto_present and bool(_chunks) and len(_chunks[-1]) <= 2

                        _render_chunks = _chunks[:-1] if _merge_panto else _chunks
                        for chunk in _render_chunks:
                            _cols = st.columns(4)
                            for col, n in zip(_cols, chunk):
                                checked = col.checkbox(n, value=(n in sel), key=f"v3_nut_{n}")
                                if checked:
                                    sel.add(n)
                                else:
                                    sel.discard(n)

                        if _merge_panto:
                            # Last row: remaining items (1–2) each get 1 unit; Pantothensäure gets 2
                            _last = _chunks[-1]
                            _widths = [1] * len(_last) + [2]
                            _cols = st.columns(_widths)
                            for col, n in zip(_cols, _last):
                                checked = col.checkbox(n, value=(n in sel), key=f"v3_nut_{n}")
                                if checked:
                                    sel.add(n)
                                else:
                                    sel.discard(n)
                            checked = _cols[len(_last)].checkbox(
                                _panto, value=(_panto in sel), key=f"v3_nut_{_panto}"
                            )
                            if checked:
                                sel.add(_panto)
                            else:
                                sel.discard(_panto)
                        elif _panto_present:
                            # No room on last row — own full-width row
                            checked = st.checkbox(
                                _panto, value=(_panto in sel), key=f"v3_nut_{_panto}"
                            )
                            if checked:
                                sel.add(_panto)
                            else:
                                sel.discard(_panto)

                        if extra_in_data:
                            st.markdown(
                                "<hr style='border:0;border-top:1px solid rgba(49,51,63,0.15);"
                                "margin:0.8rem 0 0.6rem 0;'>"
                                "<div style='font-size:1.05rem;font-weight:700;color:#19425e;"
                                "margin-bottom:0.4rem;'>➕ Additional nutrients</div>",
                                unsafe_allow_html=True,
                            )
                            for chunk in [extra_in_data[i:i+3] for i in range(0, len(extra_in_data), 3)]:
                                for col, n in zip(st.columns(len(chunk)), chunk):
                                    checked = col.checkbox(n, value=(n in sel), key=f"v3_nut_{n}")
                                    if checked:
                                        sel.add(n)
                                    else:
                                        sel.discard(n)

                    st.session_state["v3_nutrient_selection"] = sel
                    selected_nutrients  = [n for n in nutrient_names if n in sel]
                    _ft_needs_review: list = []   # populated inside else block below
                    _ft_map_locked: bool  = bool(st.session_state.get("v3_ft_map_locked", False))

                    if not selected_nutrients:
                        st.warning("⚠️ No nutrients selected.")
                    else:
                        st.caption(f"{len(selected_nutrients)} nutrients selected.")

                        # ── Futtermittel Column Mapping ───────────────
                        _ft_cols_all = get_futtermittel_column_names(file_bytes)
                        _ft_map_sig  = tuple(selected_nutrients)
                        if (st.session_state.get("v3_ft_map") is None
                                or st.session_state.get("v3_ft_map_sig") != _ft_map_sig):
                            _ftm, _fts = build_futtermittel_nutrient_mapping(
                                selected_nutrients, _ft_cols_all
                            )
                            st.session_state["v3_ft_map"]        = _ftm
                            st.session_state["v3_ft_map_sig"]    = _ft_map_sig
                            st.session_state["v3_ft_map_status"] = _fts

                        _ft_map    = st.session_state.get("v3_ft_map", {})
                        _ft_status = st.session_state.get("v3_ft_map_status", {})

                        _ft_needs_review = [r for r in selected_nutrients
                                            if _ft_status.get(r) in {"base", "fuzzy", "missing"}]
                        _ft_truly_missing = [r for r in _ft_needs_review
                                             if _ft_status.get(r) == "missing"]
                        _ft_icon = "⚠️" if _ft_truly_missing else ("✅" if _ft_map_locked else "🟡")

                        def _ft_safe_key(s: str) -> str:
                            return re.sub(r"[^a-zA-Z0-9_]", "_", s)[:80]

                        def _on_ft_sel(r):
                            _ch = st.session_state.get(f"ft_map_sel_{_ft_safe_key(r)}")
                            _m  = dict(st.session_state.get("v3_ft_map") or {})
                            _ms = dict(st.session_state.get("v3_ft_map_status") or {})
                            if _ch == "(ignore)":
                                _m[r] = None; _ms[r] = "missing"
                            else:
                                _m[r] = _ch; _ms[r] = "manual"
                            st.session_state["v3_ft_map"]        = _m
                            st.session_state["v3_ft_map_status"] = _ms

                        with st.expander(
                            f"🗺️ Futtermittel Column Mapping {_ft_icon}",
                            expanded=bool(_ft_truly_missing) or (not _ft_map_locked and bool(_ft_needs_review)),
                        ):
                            _status_icons = {
                                "exact": "✅", "base": "🟡", "fuzzy": "⚠️",
                                "missing": "❌", "manual": "🔵",
                            }
                            _ft_rows = []
                            for _r in selected_nutrients:
                                _v = _ft_map.get(_r)
                                _s = _ft_status.get(_r, "missing")
                                _ft_rows.append({
                                    "Requirement":           _r,
                                    "→ Futtermittel Column": _v or "—",
                                    "Status": f"{_status_icons.get(_s, '❓')} {_s}",
                                })
                            st.dataframe(
                                pd.DataFrame(_ft_rows),
                                use_container_width=True, hide_index=True,
                            )

                            if _ft_needs_review and not _ft_map_locked:
                                st.markdown(
                                    '<p style="margin:0.2rem 0 0.2rem 0;font-weight:600">'
                                    'Review non-exact matches:</p>',
                                    unsafe_allow_html=True,
                                )
                                for _r in _ft_needs_review:
                                    _curr = _ft_map.get(_r)
                                    _opts = ["(ignore)"] + _ft_cols_all
                                    _idx  = _opts.index(_curr) if _curr in _opts else 0
                                    st.selectbox(
                                        _r, options=_opts, index=_idx,
                                        key=f"ft_map_sel_{_ft_safe_key(_r)}",
                                        on_change=_on_ft_sel, args=(_r,),
                                    )

                            st.markdown('<span id="ft-sm-btn-marker"></span>', unsafe_allow_html=True)
                            if _ft_truly_missing:
                                st.markdown(
                                    "<div class='warnrow'>⚠️ Some nutrients have no Futtermittel match — assign or ignore them above.</div>",
                                    unsafe_allow_html=True,
                                )
                            elif _ft_map_locked:
                                st.markdown(
                                    "<div class='okrow'>✅ Column mapping locked.</div>",
                                    unsafe_allow_html=True,
                                )
                                st.markdown("<br>", unsafe_allow_html=True)
                                st.button(
                                    "🔓 Unlock & edit mapping", key="v3_ft_unlock", type="secondary",
                                    on_click=lambda: st.session_state.update({"v3_ft_map_locked": False}),
                                )
                            else:
                                st.button(
                                    "🔒️ Lock Mapping", key="v3_ft_lock", type="secondary",
                                    on_click=lambda: st.session_state.update({"v3_ft_map_locked": True}),
                                )

                        # ── Max method ───────────────────────────────
                        st.markdown("#### 📊 Maximum value method")
                        max_method = st.radio(
                            "Maximum value method",
                            ["Estimated (×3 or ×5)", "From Excel (cached)"],
                            horizontal=True,
                            key="v3_max_method",
                            label_visibility="collapsed",
                        )
                        use_calc_max = (max_method == "From Excel (cached)")

                        # ── Calculate Requirements button ─────────────
                        _file_hash = hashlib.md5(file_bytes).hexdigest()[:12]
                        calc_sig = (identifier, tuple(selected_nutrients), use_calc_max, _file_hash)

                        if st.button("🔢 Calculate Requirements", type="secondary", key="v3_calc_btn",
                                     disabled=bool(_ft_truly_missing) or not _ft_map_locked or not identifier):
                            with st.spinner("Reading requirements…"):
                                try:
                                    mins_dict, maxs_dict = get_min_values_via_openpyxl(
                                        file_bytes=file_bytes,
                                        password=_VT_PASSWORD,
                                        identifier=identifier,
                                        nutrient_names=nutrient_names,
                                    )
                                    constraints_raw = build_constraints_from_mins(
                                        mins_dict, selected_nutrients,
                                        maxs_dict=maxs_dict, use_calculated_max=use_calc_max,
                                    )
                                    st.session_state["v3_constraints_raw"]      = constraints_raw
                                    st.session_state["v3_calc_sig"]             = calc_sig
                                    st.session_state["v3_mins_dict"]            = mins_dict
                                    st.session_state["v3_maxs_dict"]            = maxs_dict
                                    st.session_state["constraints_edit_df"]     = constraints_to_edit_df(constraints_raw).copy()
                                    st.session_state["constraints_locked"]      = False
                                    st.session_state["constraints_effective_df"] = None
                                    st.session_state["constraints_editor_nonce"] = \
                                        st.session_state.get("constraints_editor_nonce", 0) + 1
                                    st.session_state["nutrient_mapping"]        = None
                                    st.session_state["nutrient_mapping_status"] = None
                                    st.session_state["nutrient_mapping_mode"]   = None
                                    st.session_state["nutrient_mapping_signature"] = None
                                    st.session_state["mapping_locked"]          = False
                                    _wt = f" | Weight: {params['weight']:.2f} kg" if params.get("weight") else ""
                                    st.success(f"✅ Requirements loaded for: **{identifier}**{_wt}")
                                except Exception as e:
                                    st.error(f"❌ Failed to read requirements: {e}")

                        elif st.session_state.get("v3_calc_sig") == calc_sig:
                            constraints_raw = st.session_state.get("v3_constraints_raw")

                        # ── Nutrient Intervals: read-only table + edit panel ──
                        if "constraints_edit_df" in st.session_state and constraints_raw is not None:
                            st.markdown("#### 🧾 Nutrient Intervals")
                            locked   = bool(st.session_state.get("constraints_locked", False))
                            _nu_intv = get_futtermittel_nutrient_units(file_bytes)
                            _cdf_now = st.session_state["constraints_edit_df"]

                            # Validate intervals
                            _min_s = pd.to_numeric(_cdf_now["Tagesbedarf (Min)"], errors="coerce").fillna(0.0)
                            _max_s = pd.to_numeric(_cdf_now["Maximalwert (Max)"], errors="coerce")
                            err_min_gt_max = _max_s.notna() & (_min_s > _max_s)
                            has_interval_errors = bool(err_min_gt_max.any())

                            # ── Read-only overview table ──────────────
                            _intv_view = pd.DataFrame({
                                "Nutrient": _cdf_now["Nährstoff"].astype(str).str.strip(),
                                "Unit":     _cdf_now["Nährstoff"].map(lambda n: _nu_intv.get(str(n).strip(), "")),
                                "Min":      _min_s.round(4),
                                "Max":      _max_s.round(4),
                            })
                            st.dataframe(
                                _intv_view,
                                use_container_width=True,
                                hide_index=True,
                                column_config={
                                    "Nutrient": st.column_config.TextColumn("Nutrient", width="medium"),
                                    "Unit":     st.column_config.TextColumn("Unit", width="small"),
                                    "Min":      st.column_config.NumberColumn("Min", format="%.4g", width="small"),
                                    "Max":      st.column_config.NumberColumn("Max", format="%.4g", width="small"),
                                },
                            )

                            if has_interval_errors:
                                st.error("❌ Invalid intervals (Min > Max). Fix before locking.")

                            # ── Edit a nutrient ───────────────────────
                            with st.expander("✏️ Edit a nutrient", expanded=False):
                                _nut_names = _cdf_now["Nährstoff"].astype(str).str.strip().tolist()
                                _edit_sel = st.selectbox(
                                    "Select nutrient to edit", options=_nut_names,
                                    key="v3_edit_sel", disabled=locked,
                                )
                                if _edit_sel:
                                    _ei = _cdf_now[_cdf_now["Nährstoff"].astype(str).str.strip() == _edit_sel].index
                                    if len(_ei):
                                        _ei = _ei[0]
                                        _cur_min = float(_cdf_now.at[_ei, "Tagesbedarf (Min)"] or 0.0)
                                        _cur_max = _cdf_now.at[_ei, "Maximalwert (Max)"]
                                        _cur_max_f = float(_cur_max) if pd.notna(_cur_max) else 0.0
                                        ec1, ec2 = st.columns(2)
                                        with ec1:
                                            _new_min = st.number_input(
                                                "Min", value=_cur_min, step=0.01, format="%.4f",
                                                key=f"v3_edit_min_{_edit_sel}", disabled=locked,
                                            )
                                        with ec2:
                                            _new_max = st.number_input(
                                                "Max (0 = no max)", value=_cur_max_f, step=0.01, format="%.4f",
                                                key=f"v3_edit_max_{_edit_sel}", disabled=locked,
                                            )
                                        if st.button("💾 Update", type="secondary", key="v3_edit_upd_btn", disabled=locked):
                                            if _new_max > 0 and _new_min > _new_max:
                                                st.error("❌ Min must not exceed Max.")
                                            else:
                                                _df_upd = st.session_state["constraints_edit_df"].copy()
                                                _df_upd.at[_ei, "Tagesbedarf (Min)"] = float(_new_min)
                                                _df_upd.at[_ei, "Maximalwert (Max)"] = float(_new_max) if _new_max > 0 else np.nan
                                                _df_upd.at[_ei, "Bedarf nach Grundnahrung (Min-Base)"] = max(
                                                    0.0, float(_new_min) - float(_df_upd.at[_ei, "Grundnahrung"] or 0.0)
                                                )
                                                st.session_state["constraints_edit_df"] = _df_upd
                                                st.session_state["nutrient_mapping"] = None
                                                st.session_state["mapping_locked"]  = False
                                                st.success(f"✅ '{_edit_sel}' updated.")
                                                st.rerun()

                            # ── Delete nutrients ──────────────────────
                            with st.expander("🗑️ Delete nutrients", expanded=False):
                                _del_opts = _cdf_now["Nährstoff"].astype(str).str.strip().tolist()
                                _del_sel  = st.multiselect(
                                    "Select nutrients to delete", options=_del_opts,
                                    key="v3_del_ms", placeholder="Search and select…",
                                    disabled=locked,
                                )
                                if _del_sel:
                                    st.warning(f"⚠️ {len(_del_sel)} nutrient(s) selected for deletion.")
                                if st.button("🗑 Delete selected", type="secondary",
                                             key="v3_delete_rows_btn", disabled=locked or not _del_sel):
                                    _df_del = st.session_state["constraints_edit_df"].copy()
                                    _df_del = _df_del[~_df_del["Nährstoff"].astype(str).str.strip().isin(set(_del_sel))]
                                    _df_del = _df_del.reset_index(drop=True)
                                    st.session_state["constraints_edit_df"] = _df_del
                                    st.session_state["nutrient_mapping"] = None
                                    st.session_state["mapping_locked"]  = False
                                    st.success(f"✅ {len(_del_sel)} nutrient(s) deleted.")
                                    st.rerun()

                            # ── Add new nutrient ──────────────────────
                            with st.expander("➕ Add new nutrient", expanded=False):
                                add_c1, add_c2, add_c3 = st.columns([2.2, 1.2, 1.2])
                                with add_c1:
                                    new_name = st.text_input("Nutrient name", value="", disabled=locked, key="v3_new_name")
                                with add_c2:
                                    new_min_val = st.number_input("Min", value=0.0, step=0.01, format="%.2f", disabled=locked, key="v3_new_min")
                                with add_c3:
                                    new_max_val = st.number_input("Max", value=0.0, step=0.01, format="%.2f", disabled=locked, key="v3_new_max")
                                new_base_val = st.number_input("Base Diet", value=0.0, step=0.01, format="%.2f", disabled=locked, key="v3_new_base")
                                if st.button("➕ Add", type="secondary", key="v3_add_btn", disabled=locked):
                                    nc = (new_name or "").strip()
                                    if not nc:
                                        st.error("Please enter a nutrient name.")
                                    else:
                                        existing = set(st.session_state["constraints_edit_df"]["Nährstoff"].astype(str).str.strip())
                                        if nc in existing:
                                            st.warning("Nutrient already exists.")
                                        elif float(new_max_val) > 0 and float(new_min_val) > float(new_max_val):
                                            st.error("Min must not exceed Max.")
                                        else:
                                            new_row = pd.DataFrame([{
                                                "Nährstoff": nc,
                                                "Tagesbedarf (Min)": float(new_min_val),
                                                "Maximalwert (Max)": float(new_max_val) if new_max_val else np.nan,
                                                "Grundnahrung": float(new_base_val),
                                                "Bedarf nach Grundnahrung (Min-Base)": max(0.0, float(new_min_val) - float(new_base_val)),
                                                "🗑 Löschen": False, "⚠️ Fehler": "",
                                            }])
                                            st.session_state["constraints_edit_df"] = pd.concat(
                                                [st.session_state["constraints_edit_df"], new_row], ignore_index=True
                                            )
                                            st.session_state["nutrient_mapping"] = None
                                            st.session_state["mapping_locked"]  = False
                                            st.success(f"✅ '{nc}' added.")
                                            st.rerun()

                            # ── Reset + Lock ──────────────────────────
                            if st.button("↩️ Reset all changes", type="secondary", key="v3_reset_btn", disabled=locked):
                                cr = st.session_state.get("v3_constraints_raw")
                                if cr is not None:
                                    st.session_state["constraints_edit_df"] = constraints_to_edit_df(cr).copy()
                                st.session_state["constraints_locked"]       = False
                                st.session_state["constraints_effective_df"] = None
                                st.session_state["nutrient_mapping"]         = None
                                st.session_state["mapping_locked"]           = False
                                st.success("Reset.")
                                st.rerun()
                            st.markdown(
                                "<hr style='border:0;border-top:1px solid rgba(49,51,63,0.22);"
                                "margin:0 0 0.6rem 0;padding:0;'/>",
                                unsafe_allow_html=True,
                            )
                            _is_locked = st.session_state.get("constraints_locked", False)
                            _, _fix_col, _ = st.columns([1, 2, 1])
                            with _fix_col:
                                if st.button(
                                    "🔓 Unlock intervals" if _is_locked else "🔏 Fix nutrient intervals",
                                    type="secondary",
                                    key="v3_fix_ni_btn",
                                    use_container_width=True,
                                ):
                                    st.session_state["constraints_locked"] = not _is_locked
                                    st.rerun()

                            if st.session_state.get("constraints_locked", False):
                                if has_interval_errors:
                                    st.session_state["constraints_locked"]      = False
                                    st.session_state["constraints_effective_df"] = None
                                    st.error("❌ Cannot fix: invalid intervals detected.")
                                    st.rerun()
                                names = st.session_state["constraints_edit_df"]["Nährstoff"].astype(str).str.strip()
                                if (names == "").any():
                                    st.session_state["constraints_locked"] = False
                                    st.error("❌ Cannot fix: empty nutrient names.")
                                    st.rerun()
                                dup_mask = names.duplicated(keep=False)
                                if dup_mask.any():
                                    dups = sorted(set(names[dup_mask].tolist()))
                                    st.session_state["constraints_locked"] = False
                                    st.error(f"❌ Duplicate nutrient names: {', '.join(dups)}")
                                    st.rerun()
                                st.session_state["constraints_effective_df"] = edit_df_to_constraints(
                                    st.session_state["constraints_edit_df"].drop(
                                        columns=["🗑 Löschen", "⚠️ Fehler"], errors="ignore"
                                    ),
                                    constraints_raw,
                                )
                                constraints_effective = st.session_state["constraints_effective_df"]
                                st.markdown("<div class='okrow' style='margin-bottom:1rem;'>✅ Intervals fixed — ready for optimization.</div>",
                                            unsafe_allow_html=True)
                            else:
                                constraints_effective = None
                                st.markdown("<div class='warnrow warnrow-after'>⏳ Not fixed yet — optimization disabled.</div>",
                                            unsafe_allow_html=True)

                    volle_tabelle_ok = _ft_map_locked and (not _ft_truly_missing) and constraints_raw is not None and constraints_effective is not None

                except Exception as e:
                    st.error("❌ Could not load volle_Tabelle.xlsm.")
                    st.caption(f"Technical notice: {e}")

    # --------------------------------------------------------
    # RIGHT: Supplement database
    # --------------------------------------------------------
    with right:
        with st.expander("Supplement Database", expanded=True):
            st.markdown("<div class='upload-title' style='margin-bottom:0.6rem;'>Supplement database</div>", unsafe_allow_html=True)

            supp_file = st.file_uploader(
                "Supplement DB Upload", type="xlsx",
                label_visibility="collapsed", key="supp_uploader",
            )
            proceed_without_incomplete = False

            if supp_file:
                try:
                    supplements = parse_supplements_excel(supp_file)
                    issues = validate_supplements_df(supplements)
                    excluded_supps_table = get_excluded_supplements(supplements)

                    if issues:
                        st.warning("⚠️ **Warning:**\n\n" + "\n".join(f"- {m}" for m in issues))

                    if not excluded_supps_table.empty:
                        st.info(f"{len(excluded_supps_table)} supplements have no valid price and will be excluded.")
                        with st.expander(f"Show details ({len(excluded_supps_table)} rows)"):
                            st.dataframe(excluded_supps_table, use_container_width=True, hide_index=True)
                        proceed_without_incomplete = st.checkbox(
                            "⏭️ Continue without incomplete supplements.", value=False, key="v3_proceed_cb"
                        )

                    supp_ok = (supplements is not None) and (proceed_without_incomplete or excluded_supps_table.empty)

                    if supp_ok:
                        st.markdown("<div class='okrow' style='margin-bottom:1rem;'>✅ Format looks good!</div>", unsafe_allow_html=True)

                        # ── Browse & Exclude ──────────────────────────────────
                        with st.expander("📋 Browse & Exclude Supplements", expanded=False):
                            # Search input
                            _sb_search = st.text_input(
                                "Search by name",
                                placeholder="Type to filter…",
                                key="v3_supp_browse_search",
                                label_visibility="collapsed",
                            )

                            # Build display table
                            _sb_price = pd.to_numeric(
                                supplements["Preis (€) pro kg"], errors="coerce"
                            )
                            _sb_excl_now = set(st.session_state.get("v3_excl_ms", []))
                            _sb_rows = supplements[["Futtermittel"]].copy()
                            _sb_rows["Price (€/kg)"] = _sb_price.round(2)
                            _sb_rows["Excluded"] = _sb_rows["Futtermittel"].isin(_sb_excl_now)
                            _sb_rows = _sb_rows.rename(columns={"Futtermittel": "Name"})
                            _sb_rows = _sb_rows.dropna(subset=["Name"])

                            if _sb_search.strip():
                                _sb_filtered = _sb_rows[
                                    _sb_rows["Name"].str.contains(_sb_search.strip(), case=False, na=False)
                                ]
                            else:
                                _sb_filtered = _sb_rows

                            st.caption(f"Showing {len(_sb_filtered)} of {len(_sb_rows)} supplements")
                            st.dataframe(
                                _sb_filtered.reset_index(drop=True),
                                use_container_width=True,
                                hide_index=True,
                                column_config={
                                    "Name":         st.column_config.TextColumn("Name",         width="large"),
                                    "Price (€/kg)": st.column_config.NumberColumn("Price (€/kg)", format="%.2f", width="small"),
                                    "Excluded":     st.column_config.CheckboxColumn("Excluded",  width="small"),
                                },
                            )

                            # Exclusion multiselect
                            _sb_all_names = sorted(
                                supplements["Futtermittel"].dropna().unique().tolist()
                            )
                            _excl_selected = st.multiselect(
                                "Exclude from optimization",
                                options=_sb_all_names,
                                default=list(st.session_state.get("v3_excl_ms", [])),
                                key="v3_excl_ms",
                                placeholder="Select supplements to exclude…",
                            )
                            if _excl_selected:
                                st.warning(
                                    f"⚠️ {len(_excl_selected)} supplement(s) excluded — "
                                    f"they will not appear in optimization results."
                                )

                    else:
                        st.markdown("<div class='warnrow'>⏳ Tick the checkbox to continue.</div>", unsafe_allow_html=True)

                except Exception as e:
                    st.error("❌ Could not parse supplement database.")
                    st.caption(f"Technical notice: {e}")
                    supp_ok = False

    status_ok   = volle_tabelle_ok and supp_ok
    status_icon = "✅" if status_ok else "⏳"
    st.markdown(
        f"<div class='statusline' style='font-size:1.8rem;text-align:center;'>Data Format "
        f"<span style='font-size:1.8rem;margin-left:0.2rem;'>{status_icon}</span></div>",
        unsafe_allow_html=True,
    )

# ============================================================
# 2b) GRUNDNAHRUNG (BASE DIET)
# ============================================================
_base_diet_exceeds_max: list = []

st.markdown("### 🥩 Base Diet")

with st.expander("🥩 Base Diet Selection (expand)", expanded=True):
    _fb = st.session_state.get("v3_file_bytes")
    if not _fb:
        st.info("⬆️ Upload CarniDiet© file first to enable base diet selection.")
    else:
        _fdf = parse_futtermittel_sheet(_fb)
        if _fdf.empty:
            st.warning("Could not load Futtermittel data from volle_Tabelle.xlsm.")
        else:
            _FT_META = {"Identifier", "Typ", "Spezies", "Kategorie", "Num", "Name", "Preis (€/kg)"}
            _ft_nutr_cols = [c for c in _fdf.columns if c not in _FT_META]

            # Species from main app settings — drives food list for non-Einzelfuttermittel
            _app_spezies = st.session_state.get("v3_animal", "Hund")

            if "v3_base_diet" not in st.session_state:
                st.session_state["v3_base_diet"] = []

            # Re-sync basket nutrient values from _fdf every render so that
            # any fix to parse_futtermittel_sheet is picked up automatically
            # without the user needing to re-add foods.
            for _bi in range(len(st.session_state["v3_base_diet"])):
                _bn = st.session_state["v3_base_diet"][_bi].get("name", "")
                _match = _fdf[_fdf["Name"] == _bn]
                if not _match.empty:
                    _fr = _match.iloc[0]
                    st.session_state["v3_base_diet"][_bi]["nutrients"] = {
                        c: float(_fr[c]) for c in _ft_nutr_cols if c in _match.columns
                    }

            gnd_left, gnd_right = st.columns([1.1, 0.9])

            # ── LEFT: filter + add ────────────────────────────────
            with gnd_left:
                st.markdown("#### 🔍 Filter & Select Food")

                # ── Typ ────────────────────────────────────────────
                _typs_avail = sorted(_fdf["Typ"].dropna().unique())
                _sel_typ = st.selectbox("Type", ["(All)"] + _typs_avail, key="v3_gnd_typ")

                # Einzelfuttermittel = raw ingredients, species-agnostic → no species filter.
                # All other Typs (Ergänzungs-, Allein-, Diätfuttermittel) are species-specific.
                if _sel_typ == "(All)" or _sel_typ == "Einzelfuttermittel":
                    _df_typ = _fdf if _sel_typ == "(All)" else _fdf[_fdf["Typ"] == _sel_typ]
                else:
                    _df_typ = _fdf[(_fdf["Typ"] == _sel_typ) & (_fdf["Spezies"] == _app_spezies)]

                # ── Kategorie ──────────────────────────────────────
                # Valid categories per Typ as defined by Combobox-Quellen sheet.
                # Items mis-tagged outside these categories are excluded.
                _VALID_KATS: dict = {
                    "Einzelfuttermittel":      {"Fleisch & Fisch", "Innereien, TNPs, Schlachtabfälle & Tiermehle",
                                                "Ei- & Milchprodukte", "Getreide, Samen & Kartoffeln",
                                                "Gemüse, Obst & Strukturtoffe", "Fette & Öle", "Mineralstoffe"},
                    "Alleinfuttermittel":      {"Feuchtfutter", "Halbfeuchte Futter", "Trockenfutter"},
                    "Ergänzungsfuttermittel":  {"vitaminierte Mineralfutter", "Vitamin/Spurenelement-Präparat",
                                                "sonst. Supplemente"},
                    "Diätfuttermittel":        {"Feuchtfutter", "Trockenfutter"},
                }
                _valid_set = _VALID_KATS.get(_sel_typ, None)
                _raw_kats  = _df_typ["Kategorie"].dropna().unique()
                if _valid_set is not None:
                    _kats_avail = sorted(k for k in _raw_kats if k in _valid_set)
                    _df_typ = _df_typ[_df_typ["Kategorie"].isin(_valid_set)]
                else:
                    _kats_avail = sorted(_raw_kats)
                # Validate stored Kategorie against current options; reset if stale
                if st.session_state.get("v3_gnd_kat", "(All)") not in ["(All)"] + _kats_avail:
                    st.session_state["v3_gnd_kat"] = "(All)"
                _sel_kat = st.selectbox("Category", ["(All)"] + _kats_avail, key="v3_gnd_kat")
                _df_flt = _df_typ if _sel_kat == "(All)" else _df_typ[_df_typ["Kategorie"] == _sel_kat]

                # ── Reset product multiselect when filter changes ──
                _flt_sig = (_sel_typ, _sel_kat)
                if st.session_state.get("v3_gnd_flt_sig") != _flt_sig:
                    st.session_state["v3_gnd_products"] = []
                    st.session_state["v3_gnd_flt_sig"] = _flt_sig

                st.caption(f"{len(_df_flt)} food item(s) in this selection.")

                _prod_opts = _df_flt["Name"].tolist()
                _sel_prods = st.multiselect(
                    "Select food items", options=_prod_opts,
                    key="v3_gnd_products", placeholder="Choose one or more…",
                )

                _pg: dict = {}
                if _sel_prods:
                    st.markdown("**Grams per day:**")
                    for _pn in _sel_prods:
                        _pg[_pn] = st.number_input(
                            _pn[:40], min_value=0.0, max_value=5000.0,
                            value=100.0, step=1.0, format="%.0f",
                            key=f"v3_gnd_g_{_pn}",
                        )

                    if st.button("➕ Add to base diet", type="secondary", key="v3_gnd_add"):
                        _bsk = list(st.session_state["v3_base_diet"])
                        for _pn in _sel_prods:
                            _pr = _df_flt[_df_flt["Name"] == _pn].iloc[0]
                            _bsk.append({
                                "name": _pn,
                                "num": int(_pr["Num"]),
                                "identifier": str(_pr["Identifier"]),
                                "grams": float(_pg.get(_pn, 100.0)),
                                "nutrients": {c: float(_pr[c]) for c in _ft_nutr_cols},
                            })
                        st.session_state["v3_base_diet"] = _bsk
                        st.success(f"✅ Added {len(_sel_prods)} item(s) to base diet.")
                        st.rerun()

                # ── Search by # ────────────────────────────────────
                st.markdown("**Add base diet by #:**")
                _num_col, _g_col = st.columns([1, 1])
                with _num_col:
                    _search_num = st.number_input(
                        "#", min_value=1, max_value=9999, value=1,
                        step=1, format="%d", key="v3_gnd_search_num",
                        label_visibility="collapsed",
                    )
                _num_match = _fdf[_fdf["Num"] == int(_search_num)]
                if not _num_match.empty:
                    _nm_row = _num_match.iloc[0]
                    st.caption(f"#{int(_search_num)}: **{_nm_row['Name']}** "
                               f"({_nm_row['Typ']} / {_nm_row['Kategorie']})")
                    with _g_col:
                        _num_grams = st.number_input(
                            "g/day", min_value=0, max_value=5000,
                            value=100, step=1, format="%d",
                            key="v3_gnd_num_grams",
                            label_visibility="collapsed",
                        )
                    if st.button("➕ Add #", type="secondary", key="v3_gnd_add_num"):
                        _bsk = list(st.session_state["v3_base_diet"])
                        _pn = _nm_row["Name"]
                        _bsk.append({
                            "name": _pn,
                            "num": int(_nm_row["Num"]),
                            "identifier": str(_nm_row["Identifier"]),
                            "grams": float(_num_grams),
                            "nutrients": {c: float(_nm_row[c]) for c in _ft_nutr_cols},
                        })
                        st.session_state["v3_base_diet"] = _bsk
                        st.success(f"✅ Added {_pn} ({int(_num_grams)}g).")
                        st.rerun()
                else:
                    st.caption(f"No food item found for #{int(_search_num)}.")

            # ── RIGHT: basket display ─────────────────────────────
            with gnd_right:
                st.markdown("#### 🧺 Current Base Diet")
                _bsk = list(st.session_state["v3_base_diet"])
                if not _bsk:
                    st.info("No items added yet.")
                else:
                    _rmv = []
                    for _bi, _bitem in enumerate(_bsk):
                        _bc1, _bc2, _bc3 = st.columns([3, 1.5, 0.5])
                        with _bc1:
                            st.markdown(f"**{_bitem['name']}**")
                        with _bc2:
                            _ng = st.number_input(
                                "g/day", min_value=0.0, max_value=5000.0,
                                value=float(_bitem["grams"]),
                                step=1.0, format="%.0f",
                                key=f"v3_bsk_g_{_bi}",
                                label_visibility="collapsed",
                            )
                            _bsk[_bi]["grams"] = _ng
                        with _bc3:
                            if st.button("🗑", key=f"v3_bsk_del_{_bi}", help="Remove"):
                                _rmv.append(_bi)

                    if _rmv:
                        _bsk = [x for j, x in enumerate(_bsk) if j not in _rmv]
                        st.session_state["v3_base_diet"] = _bsk
                        st.rerun()
                    else:
                        st.session_state["v3_base_diet"] = _bsk

                    if st.button("🗑 Clear all", type="secondary", key="v3_bsk_clear"):
                        st.session_state["v3_base_diet"] = []
                        st.rerun()

            # ── Compute base diet totals per nutrient ─────────────
            _bsk = st.session_state["v3_base_diet"]
            _base_totals: dict = {}
            for _bitem in _bsk:
                _g = float(_bitem.get("grams", 0.0))
                for _nc, _nv in _bitem.get("nutrients", {}).items():
                    _base_totals[_nc] = _base_totals.get(_nc, 0.0) + float(_nv) * _g / 100.0
            _ft_map_now = st.session_state.get("v3_ft_map", {})

            # ── Sync Grundnahrung column in constraints_edit_df ───
            _cdf_ed = st.session_state.get("constraints_edit_df")
            if _cdf_ed is not None:
                _cdf_up = _cdf_ed.copy().reset_index(drop=True)
                for _ci in range(len(_cdf_up)):
                    _rn = str(_cdf_up.at[_ci, "Nährstoff"])
                    _cdf_up.at[_ci, "Grundnahrung"] = _lookup_base_total(_base_totals, _rn, ft_mapping=_ft_map_now)
                _cdf_up["Bedarf nach Grundnahrung (Min-Base)"] = (
                    _cdf_up["Tagesbedarf (Min)"] - _cdf_up["Grundnahrung"]
                ).clip(lower=0.0)
                st.session_state["constraints_edit_df"] = _cdf_up
                # Rebuild constraints_effective if intervals are locked
                if st.session_state.get("constraints_locked", False):
                    _cr = st.session_state.get("v3_constraints_raw")
                    if _cr is not None:
                        st.session_state["constraints_effective_df"] = edit_df_to_constraints(
                            _cdf_up.drop(columns=["🗑 Löschen", "⚠️ Fehler"], errors="ignore"),
                            _cr,
                        )

            # ── Dynamic coverage display ──────────────────────────
            # Pre-compute coverage rows (shared by both expanders below)
            _n_names_cov = get_nutrient_names_from_bedarf(_fb)
            _sel_set_cov = st.session_state.get("v3_nutrient_selection", set())
            _sel_nts_cov = [n for n in _n_names_cov if n in _sel_set_cov]
            _cdf_now = st.session_state.get("constraints_edit_df")
            _nu_cov = get_futtermittel_nutrient_units(_fb)
            _cov_rows = []
            if _sel_nts_cov and _cdf_now is not None:
                for _n in _sel_nts_cov:
                    _nrow = _cdf_now[_cdf_now["Nährstoff"].astype(str).str.strip() == _n]
                    if _nrow.empty:
                        continue
                    _mn = float(_nrow["Tagesbedarf (Min)"].iloc[0])
                    _mx_raw = _nrow["Maximalwert (Max)"].iloc[0]
                    _mx = float(_mx_raw) if pd.notna(_mx_raw) else float("nan")
                    _bt = _lookup_base_total(_base_totals, _n, ft_mapping=_ft_map_now)
                    _pct = (_bt / _mn * 100.0) if _mn > 0 else 0.0

                    if pd.notna(_mx) and _bt > _mx + 1e-9:
                        _status = "🔴 Exceeds Max"
                    elif _mn > 0 and _bt >= _mn - 1e-9:
                        _status = "✅ ≥ Min"
                    elif _pct >= 50:
                        _status = "🔶 > 50%"
                    else:
                        _status = "⬜ < 50%"

                    _cov_rows.append({
                        "Nutrient":  _n,
                        "Unit":      _nu_cov.get(_n, ""),
                        "Min":       round(_mn, 2),
                        "Base Diet": round(_bt, 2),
                        "Max":       round(_mx, 2) if pd.notna(_mx) else None,
                        "% of Min":  round(min(_pct, 100.0), 1),
                        "Status":    _status,
                    })

            _base_diet_exceeds_max = [r["Nutrient"] for r in _cov_rows if r["Status"] == "🔴 Exceeds Max"]

            st.markdown('<hr style="margin-top:0.25rem;margin-bottom:0.5rem;border:none;border-top:1px solid #e5e7eb;">', unsafe_allow_html=True)
            with st.expander("🔍 Food Nutrient Lookup Table (per 100g)", expanded=False):
                _nu_lu = get_futtermittel_nutrient_units(_fb)
                _lu_filter = st.text_input(
                    "Search by name or #", key="v3_lu_filter",
                    placeholder="Type to search…",
                )
                _lu_nutr_cols = [c for c in _ft_nutr_cols if c != "met. Energie"]

                if "v3_price_overrides" not in st.session_state:
                    st.session_state["v3_price_overrides"] = {}
                _price_overrides = st.session_state["v3_price_overrides"]
                _fdf_price_dict  = dict(zip(_fdf["Name"], _fdf["Preis (€/kg)"]))

                _lu_df_full = _fdf[["Num", "Name", "Typ", "Kategorie", "Preis (€/kg)"] + _lu_nutr_cols].rename(columns={"Num": "#"})
                if _lu_filter.strip():
                    _f = _lu_filter.strip().lower()
                    _mask = (
                        _lu_df_full["Name"].str.lower().str.contains(_f, na=False) |
                        _lu_df_full["#"].astype(str).str.contains(_f, na=False)
                    )
                    _lu_df_full = _lu_df_full[_mask]

                _lu_df_disp = _lu_df_full.copy()
                for _li in _lu_df_disp.index:
                    _ln = str(_lu_df_disp.at[_li, "Name"])
                    if _ln in _price_overrides:
                        _lu_df_disp.at[_li, "Preis (€/kg)"] = _price_overrides[_ln]

                n_overridden = sum(1 for n in _lu_df_disp["Name"] if str(n) in _price_overrides)
                _override_hint = f" · {n_overridden} price(s) overridden" if n_overridden else ""
                st.markdown(f"**{len(_lu_df_disp)} food item(s) shown. Values are per 100g.{_override_hint}**")
                st.caption("✏️ The Price (€/kg) column is editable — changes apply to the cost estimate below.")

                _edited_lu = st.data_editor(
                    _lu_df_disp,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "#":            st.column_config.NumberColumn("#", disabled=True, width="small", format="%d"),
                        "Name":         st.column_config.TextColumn("Name", disabled=True, width="large"),
                        "Typ":          st.column_config.TextColumn("Type", disabled=True, width="medium"),
                        "Kategorie":    st.column_config.TextColumn("Category", disabled=True, width="medium"),
                        "Preis (€/kg)": st.column_config.NumberColumn(
                            "Price (€/kg) ✏️", format="%.2f", width="small", min_value=0.0,
                        ),
                        **{
                            n: st.column_config.NumberColumn(
                                f"{n} [{_nu_lu[n]}]" if _nu_lu.get(n) else n,
                                format="%.2f", disabled=True,
                            )
                            for n in _lu_nutr_cols
                        },
                    },
                    key="v3_lu_editor",
                )

                for _li in _edited_lu.index:
                    _ln    = str(_edited_lu.at[_li, "Name"])
                    _new_p = _edited_lu.at[_li, "Preis (€/kg)"]
                    _orig_p = _fdf_price_dict.get(_ln)
                    if pd.notna(_new_p):
                        _np = float(_new_p)
                        if _orig_p is None or pd.isna(_orig_p) or abs(_np - float(_orig_p)) > 1e-9:
                            _price_overrides[_ln] = _np
                        else:
                            _price_overrides.pop(_ln, None)
                    else:
                        _price_overrides.pop(_ln, None)
                st.session_state["v3_price_overrides"] = _price_overrides

            with st.expander("📊 Nutrient Coverage by Base Diet", expanded=True):
                if not _sel_nts_cov or _cdf_now is None:
                    st.info("📌 Calculate requirements first (Upload section above) to see nutrient coverage.")
                elif _cov_rows:
                    if _base_diet_exceeds_max:
                        st.error(
                            f"🔴 Base diet exceeds the maximum for: **{', '.join(_base_diet_exceeds_max)}**"
                        )
                    _cov_df = pd.DataFrame(_cov_rows)
                    st.dataframe(
                        _cov_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Nutrient":  st.column_config.TextColumn(width="medium"),
                            "Unit":      st.column_config.TextColumn("Unit", width="small"),
                            "Min":       st.column_config.NumberColumn(format="%.2f", width="small"),
                            "Base Diet": st.column_config.NumberColumn(format="%.2f", width="small"),
                            "Max":       st.column_config.NumberColumn(format="%.2f", width="small"),
                            "% of Min":  st.column_config.ProgressColumn(
                                "Coverage Requirements (%)", min_value=0, max_value=100, format="%.0f%%",
                            ),
                            "Status":    st.column_config.TextColumn(width="small"),
                        },
                    )
                    _n_met = sum(1 for r in _cov_rows if r["Status"] == "✅ ≥ Min")
                    st.markdown(f"**{_n_met} / {len(_cov_rows)} nutrients fully covered by base diet.**")
                else:
                    st.info("No matching nutrients found in coverage table.")

            with st.expander("🔬 Nutrient Contribution Breakdown", expanded=True):
                if not _sel_nts_cov or _cdf_now is None:
                    st.info("📌 Calculate requirements first to see nutrient breakdown.")
                elif not _bsk:
                    st.info("Add food items to the base diet to see the breakdown.")
                elif _cov_rows:
                    _bk_labels = [f"{it['name']} ({int(it['grams'])}g)" for it in _bsk]
                    _nu_bk = get_futtermittel_nutrient_units(_fb)

                    _breakdown_rows = []
                    for _n in _sel_nts_cov:
                        _row = {"Nutrient": _n, "Unit": _nu_bk.get(_n, "")}
                        _total = 0.0
                        for _it, _lbl in zip(_bsk, _bk_labels):
                            _g   = float(_it.get("grams", 0.0))
                            _val = _lookup_base_total(_it.get("nutrients", {}), _n, ft_mapping=_ft_map_now) * _g / 100.0
                            _row[_lbl] = round(_val, 2)
                            _total += _val
                        _row["Total"] = round(_total, 2)
                        _breakdown_rows.append(_row)

                    if _breakdown_rows:
                        _bd_df = pd.DataFrame(_breakdown_rows)
                        _food_col_cfg = {
                            lbl: st.column_config.NumberColumn(lbl, format="%.2f")
                            for lbl in _bk_labels
                        }
                        _food_col_cfg["Nutrient"] = st.column_config.TextColumn(width="medium")
                        _food_col_cfg["Unit"]     = st.column_config.TextColumn("Unit", width="small")
                        _food_col_cfg["Total"]    = st.column_config.NumberColumn("Total", format="%.2f", width="small")
                        st.dataframe(
                            _bd_df,
                            use_container_width=True,
                            hide_index=True,
                            column_config=_food_col_cfg,
                        )
                else:
                    st.info("No matching nutrients found in coverage table.")

            with st.expander("💰 Monthly Base Diet Cost", expanded=True):
                if not _bsk:
                    st.info("Add food items to the base diet to see the cost estimate.")
                else:
                    _cost_rows = []
                    _total_day = 0.0
                    _any_missing = False
                    _cost_overrides = st.session_state.get("v3_price_overrides", {})
                    for _it in _bsk:
                        _g = float(_it.get("grams", 0.0))
                        _food_name = _it["name"]
                        if _food_name in _cost_overrides:
                            _preis = _cost_overrides[_food_name]
                        else:
                            _fmatch = _fdf[_fdf["Name"] == _food_name]
                            _preis_raw = _fmatch.iloc[0]["Preis (€/kg)"] if not _fmatch.empty else None
                            try:
                                _preis = float(_preis_raw) if _preis_raw is not None else None
                            except (TypeError, ValueError):
                                _preis = None

                        if _preis is not None:
                            _cpd = (_g / 1000.0) * _preis
                            _cpm = _cpd * 30.0
                            _total_day += _cpd
                        else:
                            _cpd = None
                            _cpm = None
                            _any_missing = True

                        _cost_rows.append({
                            "Food":           _it["name"],
                            "g/day":          _g,
                            "€/kg":           _preis,
                            "Cost/day (€)":   round(_cpd, 1) if _cpd is not None else None,
                            "Cost/month (€)": round(_cpm, 1) if _cpm is not None else None,
                        })

                    st.dataframe(
                        pd.DataFrame(_cost_rows),
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Food":           st.column_config.TextColumn("Food", width="large"),
                            "g/day":          st.column_config.NumberColumn("g/day", format="%.0f", width="small"),
                            "€/kg":           st.column_config.NumberColumn("€/kg", format="%.2f", width="small"),
                            "Cost/day (€)":   st.column_config.NumberColumn("Cost/day (€)", format="%.1f", width="small"),
                            "Cost/month (€)": st.column_config.NumberColumn("Cost/month (€)", format="%.1f", width="small"),
                        },
                    )
                    _total_month = _total_day * 30.0
                    st.markdown(
                        f"""
                        <div style="display:flex;gap:0.6rem;margin-top:0.1rem;margin-bottom:1.2rem;justify-content:center">
                          <div style="background:#ffffff;border:1px solid #e5e7eb;border-radius:10px;padding:4px 48px 6px">
                            <div style="font-size:0.85rem;color:#6b7280;line-height:1;margin-bottom:3px;text-align:center">Total / day</div>
                            <div style="font-size:1.6rem;font-weight:600;line-height:1.2;color:#111;text-align:center">€{_total_day:.1f}</div>
                          </div>
                          <div style="background:#ffffff;border:1px solid #e5e7eb;border-radius:10px;padding:4px 48px 6px">
                            <div style="font-size:0.85rem;color:#6b7280;line-height:1;margin-bottom:3px;text-align:center">Total / month</div>
                            <div style="font-size:1.6rem;font-weight:600;line-height:1.2;color:#111;text-align:center">€{_total_month:.1f}</div>
                          </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                    if _any_missing:
                        st.caption("⚠️ Some items have no price set in the database — total may be incomplete.")

            if _base_diet_exceeds_max:
                st.error(
                    f"❌ The base diet exceeds the maximum allowed value for "
                    f"**{', '.join(_base_diet_exceeds_max)}**. "
                    f"Adjust the base diet before continuing to Data Check."
                )

# ============================================================
# 3) DATA CHECK & MATCHING
# ============================================================
st.markdown("### 🛂 Data Check")

with st.expander("🛂 Data Check (expand)", expanded=True):
    if _base_diet_exceeds_max:
        st.info("❌ Adjust the base diet (maximum exceeded) before Data Check becomes available.")
    elif status_ok and ("constraints_edit_df" in st.session_state) and (supplements is not None):
        st.markdown("#### 🔁 Nutrient Matching (Requirements ↔ Supplement DB)")
        st.markdown(
            "<div class='caption-note'>Only nutrients that couldn't be matched automatically need manual assignment.</div>",
            unsafe_allow_html=True,
        )

        ration_nutrients_now = st.session_state["constraints_edit_df"]["Nährstoff"].astype(str).str.strip().tolist()
        ration_sig_now       = tuple(ration_nutrients_now)
        supp_names_sorted    = sorted(infer_available_nutrients_from_supplements(supplements))
        _fb_dc               = st.session_state.get("v3_file_bytes")
        _req_units           = get_futtermittel_nutrient_units(_fb_dc) if _fb_dc else {}

        if (st.session_state.get("nutrient_mapping") is None
                or st.session_state.get("nutrient_mapping_signature") != ration_sig_now):
            m, s, mode = build_initial_mapping_and_status(ration_nutrients_now, supp_names_sorted, req_units=_req_units)
            st.session_state["nutrient_mapping"]           = m
            st.session_state["nutrient_mapping_status"]    = s
            st.session_state["nutrient_mapping_mode"]      = mode
            st.session_state["nutrient_mapping_signature"] = ration_sig_now

        mapping        = st.session_state["nutrient_mapping"] or {}
        mapping_status = st.session_state.get("nutrient_mapping_status", {}) or {}
        mapping_mode   = st.session_state.get("nutrient_mapping_mode",   {}) or {}

        mapping_locked = bool(st.session_state.get("mapping_locked", False))

        truly_unresolved = [r for r in ration_nutrients_now
                            if mapping_status.get(r) in {"missing", "fuzzy"}]

        if mapping_locked:
            needs_manual = truly_unresolved
        else:
            needs_manual = [r for r in ration_nutrients_now
                            if mapping_status.get(r) not in {"exact"}]

        # Track ALL matched supplements (auto + manual) as unavailable for other requirements.
        all_used = set(v for v in mapping.values() if v is not None)

        def _safe_key(s: str) -> str:
            return re.sub(r"[^a-zA-Z0-9_]", "_", s)[:80]

        def _on_match_sel(r):
            choice = st.session_state.get(f"map_sel_{_safe_key(r)}")
            m  = dict(st.session_state.get("nutrient_mapping") or {})
            mo = dict(st.session_state.get("nutrient_mapping_mode") or {})
            ms = dict(st.session_state.get("nutrient_mapping_status") or {})
            if choice == "(ignore)":
                m[r] = None; mo[r] = "manual"; ms[r] = "ignored"
            else:
                m[r] = choice; mo[r] = "manual"; ms[r] = "manual"
            st.session_state["nutrient_mapping"]        = m
            st.session_state["nutrient_mapping_mode"]   = mo
            st.session_state["nutrient_mapping_status"] = ms

        if not needs_manual:
            st.markdown("<div class='okrow'>✅ All nutrients matched automatically.</div>", unsafe_allow_html=True)
        else:
            st.markdown("<div class='warnrow'>⚠️ Manual assignment needed:</div>", unsafe_allow_html=True)
            m_left, m_right = st.columns([1, 1])
            mid = int(np.ceil(len(needs_manual) / 2))

            def render_match_col(names, container):
                with container:
                    for r in names:
                        current = mapping.get(r, None)
                        # Exclude all already-matched supplements except the current assignment
                        other_used = all_used - ({current} if current else set())
                        opts = ["(ignore)"] + [sn for sn in supp_names_sorted if sn not in other_used]
                        # Build base→canonicals map so we can match on stripped names
                        # without creating duplicates when two canonicals share the same base.
                        _sugg_base_map: dict = {}
                        for x in supp_names_sorted:
                            _sugg_base_map.setdefault(normalize_name(strip_unit_brackets(x)), []).append(x)
                        sugg_bases = difflib.get_close_matches(
                            normalize_name(strip_unit_brackets(r)),
                            list(_sugg_base_map.keys()),
                            n=5, cutoff=0.60,
                        )
                        sugg_real, _seen = [], set()
                        for _base in sugg_bases:
                            for canon in _sugg_base_map.get(_base, []):
                                if canon in opts and canon not in _seen:
                                    sugg_real.append(canon); _seen.add(canon)
                        rest      = [x for x in opts[1:] if x not in _seen]
                        opts_final = ["(ignore)"] + sugg_real + rest
                        idx  = 0 if current is None else (opts_final.index(current) if current in opts_final else 0)
                        hint = (" (no match)" if mapping_status.get(r) == "missing" else
                                " (ignored)" if mapping_status.get(r) == "ignored" else
                                " (unclear)")

                        st.markdown(f"**{r}**{hint}")
                        st.selectbox(
                            "match", options=opts_final, index=idx,
                            key=f"map_sel_{_safe_key(r)}", label_visibility="collapsed",
                            on_change=_on_match_sel, args=(r,),
                        )

            render_match_col(needs_manual[:mid], m_left)
            render_match_col(needs_manual[mid:], m_right)
            st.session_state["nutrient_mapping"]        = mapping
            st.session_state["nutrient_mapping_mode"]   = mapping_mode
            st.session_state["nutrient_mapping_status"] = mapping_status

        rows = []
        for r in ration_nutrients_now:
            v = mapping.get(r, None)
            if v is None and mapping_status.get(r) == "ignored":
                st_val = "ignored"
            else:
                st_val = ("missing" if v is None else
                          (mapping_mode.get(r, "auto") if mapping_mode.get(r) == "manual" else mapping_status.get(r, "missing")))
            req_unit  = _req_units.get(r, "")
            supp_unit = extract_supp_unit(v) if v else ""
            if not req_unit or not supp_unit or req_unit == supp_unit:
                unit_flag = ""
            else:
                _factor = get_unit_conversion_factor(req_unit, supp_unit)
                unit_flag = "🔄 auto" if _factor is not None else "⚠️ incompatible"
            rows.append({
                "Ration Nutrient":   r,
                "Req. Unit":         req_unit,
                "Supplement Column": (v or ""),
                "Supp. Unit":        supp_unit,
                "Unit":              unit_flag,
                "Status":            st_val,
            })
        mapping_df = pd.DataFrame(rows)

        def compute_needed_and_status(_ce, _map, _mstatus=None):
            if _ce is None:
                return None, True, [], []
            eff_cols = list(_ce.columns)
            base_e   = pd.to_numeric(_ce.loc["Grundnahrung", eff_cols], errors="coerce").fillna(0.0)
            min_e    = pd.to_numeric(_ce.loc["Tagesbedarf",  eff_cols], errors="coerce").fillna(0.0)
            needed   = (min_e - base_e).clip(lower=0.0)
            needed_cols = [c for c in eff_cols if float(needed.get(c, 0.0)) > 0.0]
            ok, miss, skipped = True, [], []
            for n in needed_cols:
                if _map.get(n) is None:
                    if _mstatus and _mstatus.get(n) == "ignored":
                        skipped.append(n)
                    else:
                        ok = False; miss.append(n)
            return needed_cols, ok, miss, skipped

        if constraints_effective is None:
            st.markdown("<div class='warnrow'>⏳ Fix intervals first.</div>", unsafe_allow_html=True)
        else:
            _, mapping_ok, missing_needed, skipped_nutrients = compute_needed_and_status(
                constraints_effective, mapping, mapping_status
            )

            if truly_unresolved:
                st.markdown("<div class='warnrow'>⚠️ Resolve all unmatched nutrients above before fixing.</div>",
                            unsafe_allow_html=True)
            elif mapping_locked:
                st.markdown("<div class='okrow'>✅ Matching fixed — optimization can start.</div>",
                            unsafe_allow_html=True)
                if skipped_nutrients:
                    st.caption(f"Ignored (excluded from optimization): {', '.join(sorted(skipped_nutrients))}")
                st.markdown("<br>", unsafe_allow_html=True)
                st.button(
                    "🔓 Unlock & edit matching", key="v3_unlock_mapping", type="secondary",
                    on_click=lambda: st.session_state.update({"mapping_locked": False}),
                )
            else:
                if not mapping_ok:
                    st.markdown("<div class='errorrow'>❌ Required nutrients still unmatched — assign or ignore them above.</div>",
                                unsafe_allow_html=True)
                    st.json(missing_needed)
                else:
                    if skipped_nutrients:
                        st.caption(f"Will be ignored in optimization: {', '.join(sorted(skipped_nutrients))}")
                    st.button(
                        "🔒️ Fix Matching", key="v3_fix_mapping", type="secondary",
                        on_click=lambda: st.session_state.update({"mapping_locked": True}),
                    )

        with st.expander("Show full mapping table"):
            _incompat = mapping_df[mapping_df["Unit"] == "⚠️ incompatible"]
            _converted = mapping_df[mapping_df["Unit"] == "🔄 auto"]
            if not _incompat.empty:
                st.warning(
                    f"⚠️ {len(_incompat)} nutrient(s) have incompatible units that cannot be converted "
                    f"({', '.join(_incompat['Ration Nutrient'].tolist())}) — these will be compared as-is and may produce incorrect results."
                )
            if not _converted.empty:
                st.info(
                    f"🔄 {len(_converted)} nutrient(s) have convertible unit mismatches and will be automatically scaled before optimization "
                    f"({', '.join(_converted['Ration Nutrient'].tolist())})."
                )
            st.dataframe(
                mapping_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Ration Nutrient":   st.column_config.TextColumn("Requirement Nutrient", width="medium"),
                    "Req. Unit":         st.column_config.TextColumn("Req. Unit", width="small"),
                    "Supplement Column": st.column_config.TextColumn("Supplement Column", width="medium"),
                    "Supp. Unit":        st.column_config.TextColumn("Supp. Unit", width="small"),
                    "Unit":              st.column_config.TextColumn("Unit", width="small"),
                    "Status":            st.column_config.TextColumn("Status", width="small"),
                },
            )
    else:
        st.markdown("<div class='warnrow'>⏳ Upload both files and calculate requirements first.</div>",
                    unsafe_allow_html=True)

# ============================================================
# 4) OPTIMIZATION
# ============================================================
mapping_ok_for_run         = True
mapped_constraints_effective = constraints_effective

if status_ok and supplements is not None and constraints_effective is not None and "nutrient_mapping" in st.session_state:
    try:
        mapping        = st.session_state.get("nutrient_mapping") or {}
        _mstatus_run   = st.session_state.get("nutrient_mapping_status") or {}
        _fb_dc_run     = st.session_state.get("v3_file_bytes")
        _req_units_run = get_futtermittel_nutrient_units(_fb_dc_run) if _fb_dc_run else {}
        mapped_constraints_effective = apply_mapping_to_constraints(
            constraints_effective, mapping, req_units=_req_units_run
        )
        eff_cols  = list(constraints_effective.columns)
        base_e    = pd.to_numeric(constraints_effective.loc["Grundnahrung", eff_cols], errors="coerce").fillna(0.0)
        min_e     = pd.to_numeric(constraints_effective.loc["Tagesbedarf",  eff_cols], errors="coerce").fillna(0.0)
        needed    = (min_e - base_e).clip(lower=0.0)
        needed_cols = [c for c in eff_cols if float(needed.get(c, 0.0)) > 0.0]
        for n in needed_cols:
            if mapping.get(n) is None and _mstatus_run.get(n) != "ignored":
                mapping_ok_for_run = False
                break
    except Exception:
        mapping_ok_for_run = False

_mapping_locked_now = bool(st.session_state.get("mapping_locked", False))

_base_over_max_for_opt: list = []
if constraints_effective is not None:
    try:
        _eff_cols_opt = list(constraints_effective.columns)
        _base_e_opt = pd.to_numeric(constraints_effective.loc["Grundnahrung", _eff_cols_opt], errors="coerce").fillna(0.0)
        _max_e_opt  = pd.to_numeric(constraints_effective.loc["Maximaler_Wert", _eff_cols_opt], errors="coerce")
        for _c_opt in _eff_cols_opt:
            _mx_opt = _max_e_opt.get(_c_opt, float("nan"))
            if pd.notna(_mx_opt) and float(_base_e_opt.get(_c_opt, 0.0)) > float(_mx_opt) + 1e-9:
                _base_over_max_for_opt.append(_c_opt)
    except Exception:
        pass

can_run = (status_ok and _mapping_locked_now and
           (mapped_constraints_effective is not None) and
           (supplements is not None) and mapping_ok_for_run and
           not _base_over_max_for_opt)

st.markdown("### 📈 Optimization")

with st.expander("Optimization (expand)", expanded=True):
    if "opt_result" not in st.session_state:
        st.session_state["opt_result"] = None

    # ── Supplement count limit ────────────────────────────────────────────
    _use_count_limit = st.checkbox(
        "Limit number of supplements",
        value=bool(st.session_state.get("v3_use_count_limit", False)),
        key="v3_use_count_limit",
    )
    _max_supps_val = None
    if _use_count_limit:
        _max_supps_val = st.number_input(
            "Maximum supplements",
            min_value=1, max_value=50,
            value=int(st.session_state.get("v3_max_supps", 5)),
            step=1, key="v3_max_supps",
            help="The optimizer finds the cheapest combination using at most this many supplements. "
                 "Fewer may be selected if that is sufficient.",
        )

    _opt_clicked = st.button("🚀 Start Optimization", disabled=not can_run)
    if not _mapping_locked_now:
        st.info("🔒 Fix the nutrient matching before optimizing.")
    if _base_over_max_for_opt:
        st.info(
            f"❌ Base diet exceeds the maximum for **{', '.join(_base_over_max_for_opt)}** — "
            f"adjust the base diet before optimizing."
        )
    if _opt_clicked:
        ph = st.empty()
        ph.markdown("<div class='warnrow'>⏳ Optimizing…</div>", unsafe_allow_html=True)
        try:
            _max_supps_run  = int(_max_supps_val) if _use_count_limit and _max_supps_val else None
            _excl_supps_run = list(st.session_state.get("v3_excl_ms", []))
            status, solution, cost, infeasible_msg, debug = optimize_fast(
                mapped_constraints_effective, supplements,
                max_supplements=_max_supps_run,
                excluded_supplements=_excl_supps_run,
            )
        except Exception as e:
            ph.markdown(f"<div class='errorrow'>❌ Error: {e}</div>", unsafe_allow_html=True)
            st.exception(e); st.stop()

        if status != "Optimal" or solution is None:
            ph.markdown("<div class='errorrow'>❌ No optimal solution found.</div>", unsafe_allow_html=True)
            st.markdown(f"**Solver status:** `{status}`")
            if _max_supps_run is not None:
                st.warning(
                    f"The limit of **{_max_supps_run} supplement{'s' if _max_supps_run != 1 else ''}** "
                    f"may be too restrictive to meet all nutrient requirements. "
                    f"Try raising it to {_max_supps_run + 2} or higher, or uncheck the limit."
                )
            if infeasible_msg:
                st.info("Initial assessment:"); st.json(infeasible_msg)
            diag = debug.get("diagnose") if isinstance(debug, dict) else None
            if diag:
                with st.expander("🧩 Diagnosis", expanded=True):
                    likely = diag.get("likely_problem_nutrients")
                    if isinstance(likely, pd.DataFrame) and not likely.empty:
                        st.markdown("#### ⚠️ Suspicious nutrients")
                        st.dataframe(likely[["Nutrient", "Min_After_Base_Diet", "Max_After_Base_Diet",
                                              "Max_Achievable_Rough", "Has_Positive_Source", "Flags"]],
                                     use_container_width=True, hide_index=True)
                    full = diag.get("diag_df")
                    with st.expander("All nutrients (full table)"):
                        if isinstance(full, pd.DataFrame):
                            st.dataframe(full[["Nutrient", "Min_After_Base_Diet", "Max_After_Base_Diet",
                                               "Max_Achievable_Rough", "Has_Positive_Source", "Flags"]],
                                         use_container_width=True, hide_index=True)
            st.markdown("## 🧪 Debug: Slack analysis")
            slack_df = optimize_with_slacks(mapped_constraints_effective, supplements,
                                            excluded_supplements=_excl_supps_run)
            slack_df = slack_df[(slack_df["Slack_Min"] > 1e-9) | (slack_df["Slack_Max"] > 1e-9)]
            if not slack_df.empty:
                st.dataframe(slack_df, use_container_width=True, hide_index=True)
                st.info("Large slack → likely a unit issue or incorrect Excel value.")
            else:
                st.success("Slack ≈ 0 → combination or minimum-quantity issue.")
            st.stop()

        st.session_state["opt_result"] = {
            "status": status, "solution": solution, "cost": cost,
            "debug": debug, "mapped_constraints_effective": mapped_constraints_effective,
            "max_supplements": _max_supps_run,
        }
        ph.markdown("<div class='okrow'>✅ Done.</div>", unsafe_allow_html=True)

    res = st.session_state.get("opt_result")
    if res is None:
        st.info("No results yet. Start the optimization above.")
        st.stop()

    status = res["status"]; solution = res["solution"]
    cost   = res["cost"];   debug    = res["debug"]
    mapped_constraints_effective = res["mapped_constraints_effective"]
    _res_max_supps = res.get("max_supplements")

    if st.button("♻️ Reset result"):
        st.session_state["opt_result"] = None
        st.rerun()

    _result_label = "Result"
    if _res_max_supps is not None:
        _result_label += f" (max {_res_max_supps} supplement{'s' if _res_max_supps != 1 else ''})"
    st.subheader(_result_label)

    # Build enriched per-supplement dataframe
    _sc = debug.get("supp_clean")
    _sol_rows = []
    for _s, _kg in solution.items():
        _g_day    = _kg * 1000.0
        _price_kg = float(_sc.loc[_s, "Preis (€) pro kg"]) if (_sc is not None and _s in _sc.index) else float("nan")
        _cost_day = _price_kg * _kg       if not math.isnan(_price_kg) else float("nan")
        _cost_mon = _cost_day * 30.0      if not math.isnan(_cost_day) else float("nan")
        _days_1kg = 1000.0 / _g_day       if _g_day > 0                else float("nan")
        _sol_rows.append({
            "Supplement":          _s,
            "g/day":               round(_g_day, 2),
            "€/kg":                round(_price_kg, 2)  if not math.isnan(_price_kg) else None,
            "Cost/day (€)":        round(_cost_day, 1)  if not math.isnan(_cost_day) else None,
            "Cost/month (€)":      round(_cost_mon, 1)  if not math.isnan(_cost_mon) else None,
            "Days Supply (1 kg)":  round(_days_1kg, 1)  if not math.isnan(_days_1kg) else None,
        })
    sol_df_full = (pd.DataFrame(_sol_rows)
                   .sort_values("g/day", ascending=False)
                   .reset_index(drop=True))
    display_sol = sol_df_full[["Supplement", "g/day"]].rename(columns={"g/day": "Daily Amount (g)"})

    total_cost_day   = cost
    total_cost_month = cost * 30.0

    k1, k2, k3 = st.columns(3)
    with k1:
        st.markdown(
            f"""<div class="kpi-card" style="background:rgba(46,184,92,0.12);border:1px solid rgba(46,184,92,0.30);">
            <div class="kpi-title">💰 Cost / day</div>
            <div class="kpi-value">€ {total_cost_day:.1f}</div></div>""",
            unsafe_allow_html=True)
    with k2:
        st.markdown(
            f"""<div class="kpi-card" style="background:rgba(46,184,92,0.08);border:1px solid rgba(46,184,92,0.22);">
            <div class="kpi-title">📅 Cost / month</div>
            <div class="kpi-value">€ {total_cost_month:.1f}</div></div>""",
            unsafe_allow_html=True)
    with k3:
        st.markdown(
            f"""<div class="kpi-card"><div class="kpi-title">💊 Supplements</div>
            <div class="kpi-value">{len(solution)}</div></div>""",
            unsafe_allow_html=True)

    st.caption("Minimum quantity per supplement: 1 g/day. Days Supply assumes a 1 kg purchase per supplement.")

    st.markdown("#### 💊 Recommended supplements")
    st.dataframe(
        sol_df_full,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Supplement":         st.column_config.TextColumn("Supplement", width="large"),
            "g/day":              st.column_config.NumberColumn("g/day",            format="%.2f",  width="small"),
            "€/kg":               st.column_config.NumberColumn("€/kg",             format="%.2f",  width="small"),
            "Cost/day (€)":       st.column_config.NumberColumn("Cost/day (€)",     format="%.1f",  width="small"),
            "Cost/month (€)":     st.column_config.NumberColumn("Cost/month (€)",   format="%.1f",  width="small"),
            "Days Supply (1 kg)": st.column_config.NumberColumn("Days Supply (1 kg)", format="%.0f", width="small"),
        },
    )

    # ── Nutrient Coverage Chart ──────────────────────────────────────────────
    with st.expander("📊 Nutrient Coverage after Optimization", expanded=True):
        _nc_cols   = debug.get("nutrient_cols", [])
        _nc_sc     = debug.get("supp_clean")
        if not _nc_cols or _nc_sc is None:
            st.warning("No nutrient data available for chart.")
        else:
            _nc_cdf  = mapped_constraints_effective
            _nc_base = pd.to_numeric(_nc_cdf.loc["Grundnahrung",   _nc_cols], errors="coerce").fillna(0.0)
            _nc_minv = pd.to_numeric(_nc_cdf.loc["Tagesbedarf",    _nc_cols], errors="coerce").fillna(0.0)
            _nc_maxv = pd.to_numeric(_nc_cdf.loc["Maximaler_Wert", _nc_cols], errors="coerce")
            _nc_xkg  = solution
            _nc_supps_list = list(_nc_xkg.index)

            # per-supplement contributions  {nutrient: {supplement: value}}
            _nc_by_s: dict = {}
            for _n in _nc_cols:
                _nc_by_s[_n] = {}
                for _s in _nc_supps_list:
                    _nc_by_s[_n][_s] = (
                        float(_nc_sc.loc[_s, _n]) * float(_nc_xkg[_s])
                        if (_s in _nc_sc.index and _n in _nc_sc.columns) else 0.0
                    )

            # build per-nutrient summary rows (only nutrients with min > 0)
            _ch_data: list = []
            for _n in _nc_cols:
                _mn = float(_nc_minv.get(_n, 0.0))
                if _mn <= 0:
                    continue
                _bv   = float(_nc_base.get(_n, 0.0))
                _sv   = sum(_nc_by_s[_n].values())
                _tv   = _bv + _sv
                _mx_r = _nc_maxv.get(_n)
                _mx   = float(_mx_r) if pd.notna(_mx_r) else None
                _tol  = 1e-6 * max(1.0, abs(_mn))
                _st   = ("below" if _tv + _tol < _mn else
                         "above" if (_mx is not None and _tv > _mx + _tol) else "ok")
                _unit = extract_supp_unit(_n)
                _dname = strip_unit_brackets(_n)
                _ch_data.append({
                    "name": _n, "display": _dname, "unit": _unit,
                    "min": _mn, "max": _mx, "base": _bv,
                    "supp_total": _sv, "total": _tv, "status": _st,
                    "base_pct":  _bv / _mn * 100,
                    "supp_pct":  _sv / _mn * 100,
                    "total_pct": _tv / _mn * 100,
                    "max_pct":   (_mx / _mn * 100) if _mx is not None else None,
                    "by_s":      _nc_by_s[_n],
                })

            # shared layout helper
            def _apply_chart_layout(fig, data, h_per_row=26):
                # x-axis must fit both the bars (total_pct) and the max markers (max_pct)
                _x_candidates = [d["total_pct"] for d in data]
                _x_candidates += [d["max_pct"] for d in data if d["max_pct"] is not None]
                _x_ceil = max(max(_x_candidates) * 1.08 if _x_candidates else 130, 130)
                # min — dashed orange vertical line (full height)
                fig.add_vline(
                    x=100, line_dash="dash",
                    line_color="rgba(234,137,0,0.9)", line_width=2,
                    annotation_text="Min", annotation_position="top right",
                    annotation_font=dict(color="rgba(234,137,0,1)", size=11),
                )
                # max — dashed red vertical tick per nutrient (same visual style as min)
                for _i, d in enumerate(data):
                    if d["max_pct"] is not None:
                        fig.add_shape(
                            type="line",
                            x0=d["max_pct"], x1=d["max_pct"],
                            y0=_i - 0.44, y1=_i + 0.44,
                            line=dict(color="rgba(220,38,38,0.9)", width=2, dash="dash"),
                        )
                fig.update_layout(
                    barmode="stack",
                    xaxis_title="% of minimum requirement",
                    xaxis=dict(range=[0, _x_ceil]),
                    height=max(340, len(data) * h_per_row + 90),
                    margin=dict(l=10, r=130, t=10, b=50),
                    legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0),
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    yaxis=dict(autorange="reversed"),
                )

            _STATUS_COLORS = {
                "ok":    "rgba(66,133,244,0.80)",
                "below": "rgba(220,38,38,0.80)",
                "above": "rgba(234,137,0,0.85)",
            }
            _SUPP_PALETTE = [
                "#4285f4","#ea4335","#fbbc04","#34a853","#ff6d00",
                "#46bdc6","#7b1fa2","#e91e63","#00897b","#f4511e",
                "#5c6bc0","#8d6e63","#26a69a","#ef5350","#ab47bc",
            ]

            if _ch_data:
                _dnames = [d["display"] for d in _ch_data]

                _chart_mode = st.radio(
                    "Chart view",
                    ["Absolute values", "Per supplement"],
                    horizontal=True, key="v3_chart_mode",
                )

                # shared: bold total + [min, max] label for every nutrient row
                _txt_labels = [
                    (f"<b>{d['total']:.4g} {d['unit']}</b>   [{d['min']:.4g}, {d['max']:.4g}]"
                     if d["max"] is not None else
                     f"<b>{d['total']:.4g} {d['unit']}</b>   [min {d['min']:.4g}]")
                    for d in _ch_data
                ]

                # ── Mode 1: absolute values ───────────────────────────
                if _chart_mode == "Absolute values":
                    _fig = go.Figure()
                    _fig.add_trace(go.Bar(
                        y=_dnames,
                        x=[d["base_pct"] for d in _ch_data],
                        name="Base Diet",
                        orientation="h",
                        marker_color="rgba(52,168,83,0.65)",
                        customdata=[[d["base"], d["unit"], d["min"]] for d in _ch_data],
                        hovertemplate=(
                            "<b>%{y}</b><br>"
                            "Base diet: %{customdata[0]:.4g} %{customdata[1]}<br>"
                            "(%{x:.1f}% of min = %{customdata[2]:.4g} %{customdata[1]})"
                            "<extra></extra>"
                        ),
                    ))
                    _fig.add_trace(go.Bar(
                        y=_dnames,
                        x=[d["supp_pct"] for d in _ch_data],
                        name="Supplements",
                        orientation="h",
                        marker=dict(color=[_STATUS_COLORS[d["status"]] for d in _ch_data]),
                        customdata=[[d["supp_total"], d["unit"], d["total"], d["max"]] for d in _ch_data],
                        hovertemplate=(
                            "<b>%{y}</b><br>"
                            "Supplements: %{customdata[0]:.4g} %{customdata[1]}<br>"
                            "Total: %{customdata[2]:.4g} %{customdata[1]}<br>"
                            "Max: %{customdata[3]:.4g} %{customdata[1]}"
                            "<extra></extra>"
                        ),
                    ))
                    for _ai, (_ad, _albl) in enumerate(zip(_ch_data, _txt_labels)):
                        _fig.add_annotation(
                            x=1.01, y=_ad["display"],
                            xref="paper", yref="y",
                            text=_albl, showarrow=False,
                            font=dict(size=12, color="#111"),
                            align="left", xanchor="left",
                        )
                    _apply_chart_layout(_fig, _ch_data, h_per_row=30)
                    _fig.update_layout(margin=dict(l=10, r=420, t=10, b=50))
                    st.plotly_chart(_fig, use_container_width=True)
                    st.caption(
                        "🟩 Base diet · 🟦 Supplements (ok) · 🔴 below min · 🟧 above max · "
                        "— dashed orange = Min · — dashed red = Max · "
                        "label: <b>total</b> [min, max] in original units"
                    )

                # ── Mode 2: per supplement ────────────────────────────
                else:
                    _fig = go.Figure()
                    _fig.add_trace(go.Bar(
                        y=_dnames,
                        x=[d["base_pct"] for d in _ch_data],
                        name="Base Diet",
                        orientation="h",
                        marker_color="rgba(52,168,83,0.65)",
                        customdata=[[d["base"], d["unit"]] for d in _ch_data],
                        hovertemplate=(
                            "<b>%{y}</b><br>Base diet: %{customdata[0]:.4g} %{customdata[1]}"
                            " (%{x:.1f}% of min)<extra></extra>"
                        ),
                    ))
                    for _si, _s in enumerate(_nc_supps_list):
                        _pct_vals = [
                            d["by_s"].get(_s, 0.0) / d["min"] * 100
                            for d in _ch_data
                        ]
                        _abs_vals = [d["by_s"].get(_s, 0.0) for d in _ch_data]
                        _col = _SUPP_PALETTE[_si % len(_SUPP_PALETTE)]
                        _label = (_s[:28] + "…") if len(_s) > 29 else _s
                        _fig.add_trace(go.Bar(
                            y=_dnames,
                            x=_pct_vals,
                            name=_label,
                            orientation="h",
                            marker_color=_col,
                            customdata=[[_abs_vals[j], d["unit"]] for j, d in enumerate(_ch_data)],
                            hovertemplate=(
                                f"<b>%{{y}}</b><br>{_label}: %{{customdata[0]:.4g}} %{{customdata[1]}}"
                                " (%{x:.1f}% of min)<extra></extra>"
                            ),
                        ))
                    for _ai, (_ad, _albl) in enumerate(zip(_ch_data, _txt_labels)):
                        _fig.add_annotation(
                            x=1.01, y=_ad["display"],
                            xref="paper", yref="y",
                            text=_albl, showarrow=False,
                            font=dict(size=12, color="#111"),
                            align="left", xanchor="left",
                        )
                    _apply_chart_layout(_fig, _ch_data, h_per_row=28)
                    _fig.update_layout(margin=dict(l=10, r=420, t=10, b=50))
                    st.plotly_chart(_fig, use_container_width=True)
                    st.caption(
                        "🟩 Base diet · each other color = one supplement · "
                        "— dashed orange = Min · — dashed red = Max · "
                        "label: <b>total</b> [min, max] in original units"
                    )

                # ── Absolute values table (always shown) ──────────────
                st.markdown("#### 📋 Nutrient Coverage Table")

                # Build per-food contribution columns ─────────────────
                _bsk_nc      = st.session_state.get("v3_base_diet", [])
                _mapping_nc  = st.session_state.get("nutrient_mapping") or {}
                _inv_map_nc  = {v: k for k, v in _mapping_nc.items() if v is not None}
                _ft_map_nc   = st.session_state.get("v3_ft_map") or {}
                _fb_nc       = st.session_state.get("v3_file_bytes")
                _req_u_nc    = get_futtermittel_nutrient_units(_fb_nc) if _fb_nc else {}

                def _food_col_key(name, grams, used):
                    short = (name[:18] + "…") if len(name) > 19 else name
                    key = f"{short} ({grams:.0f}g)"
                    if key not in used:
                        return key
                    i = 2
                    while f"{key}#{i}" in used:
                        i += 1
                    return f"{key}#{i}"

                _food_cols: list = []   # [(col_key, {mapped_n: value})]
                _used_food_keys: set = set()
                for _bitem in _bsk_nc:
                    _fname = _bitem.get("name", "")
                    _fg    = float(_bitem.get("grams", 0.0))
                    _fnutr = _bitem.get("nutrients", {})
                    _fkey  = _food_col_key(_fname, _fg, _used_food_keys)
                    _used_food_keys.add(_fkey)
                    _fcontrib: dict = {}
                    for _n in _nc_cols:
                        _req_n  = _inv_map_nc.get(_n)
                        _ft_col = _ft_map_nc.get(_req_n) if _req_n else None
                        if _ft_col and _ft_col in _fnutr:
                            _raw = float(_fnutr[_ft_col]) * _fg / 100.0
                            _fac = get_unit_conversion_factor(
                                _req_u_nc.get(_req_n, ""), extract_supp_unit(_n)
                            )
                            _fcontrib[_n] = _raw * (_fac if _fac is not None else 1.0)
                        else:
                            _fcontrib[_n] = 0.0
                    _food_cols.append((_fkey, _fcontrib))

                # Build supplement column keys with dose ───────────────
                _supp_col_keys: dict = {}   # supplement_name → col_key
                _used_supp_keys: set = set()
                for _s in _nc_supps_list:
                    _sdose_g = float(_nc_xkg.get(_s, 0.0)) * 1000
                    _sshort  = (_s[:16] + "…") if len(_s) > 17 else _s
                    _skey    = f"{_sshort} ({_sdose_g:.1f}g)"
                    if _skey in _used_supp_keys:
                        _i = 2
                        while f"{_skey}#{_i}" in _used_supp_keys:
                            _i += 1
                        _skey = f"{_skey}#{_i}"
                    _used_supp_keys.add(_skey)
                    _supp_col_keys[_s] = _skey

                _abs_rows = []
                for d in _ch_data:
                    _row = {
                        "Nutrient": d["display"],
                        "Unit":     d["unit"],
                        "Min":      d["min"],
                        "Max":      d["max"] if d["max"] is not None else None,
                    }
                    for _fkey, _fcontrib in _food_cols:
                        _row[_fkey] = round(_fcontrib.get(d["name"], 0.0), 6)
                    for _s in _nc_supps_list:
                        _row[_supp_col_keys[_s]] = round(d["by_s"].get(_s, 0.0), 6)
                    _row["Total"]  = round(d["total"], 6)
                    _row["Status"] = ("✅ ok" if d["status"] == "ok" else
                                      "⬇️ below min" if d["status"] == "below" else "⬆️ above max")
                    _abs_rows.append(_row)
                _abs_df = pd.DataFrame(_abs_rows)
                _col_cfg_abs = {
                    "Nutrient": st.column_config.TextColumn("Nutrient", width="medium"),
                    "Unit":     st.column_config.TextColumn("Unit",     width="small"),
                    "Min":      st.column_config.NumberColumn("Min",    format="%.4g", width="small"),
                    "Max":      st.column_config.NumberColumn("Max",    format="%.4g", width="small"),
                    "Total":    st.column_config.NumberColumn("Total",  format="%.4g", width="small"),
                    "Status":   st.column_config.TextColumn("Status",   width="small"),
                }
                for _fkey, _ in _food_cols:
                    _col_cfg_abs[_fkey] = st.column_config.NumberColumn(_fkey, format="%.4g", width="small")
                for _skey in _supp_col_keys.values():
                    _col_cfg_abs[_skey] = st.column_config.NumberColumn(_skey, format="%.4g", width="small")
                st.dataframe(_abs_df, use_container_width=True, hide_index=True,
                             column_config=_col_cfg_abs)

    with st.expander("📊 Summary Table", expanded=False):
        nutrient_cols = debug.get("nutrient_cols", [])
        supp_clean    = debug.get("supp_clean")
        if supp_clean is None or not nutrient_cols:
            st.warning("No nutrient data available.")
        else:
            x_kg        = solution
            supp_intake = {n: float((supp_clean[n].reindex(x_kg.index).fillna(0.0) * x_kg).sum())
                           for n in nutrient_cols}
            cdf  = mapped_constraints_effective
            base = pd.to_numeric(cdf.loc["Grundnahrung",   nutrient_cols], errors="coerce").fillna(0.0)
            minv = pd.to_numeric(cdf.loc["Tagesbedarf",    nutrient_cols], errors="coerce").fillna(0.0)
            maxv = pd.to_numeric(cdf.loc["Maximaler_Wert", nutrient_cols], errors="coerce")
            after = {n: float(base.get(n, 0.0)) + float(supp_intake.get(n, 0.0)) for n in nutrient_cols}
            nc_df = pd.DataFrame({
                "Nutrient":            nutrient_cols,
                "Minimum":             [float(minv.get(n, 0.0)) for n in nutrient_cols],
                "Maximum":             [float(maxv.get(n)) if pd.notna(maxv.get(n)) else np.nan for n in nutrient_cols],
                "Base Diet":           [float(base.get(n, 0.0)) for n in nutrient_cols],
                "Supplementation":     [float(supp_intake.get(n, 0.0)) for n in nutrient_cols],
                "Base Diet + Supplements": [after[n] for n in nutrient_cols],
            })
            def _status(row):
                tol    = 1e-9 + 1e-6 * max(1.0, abs(row["Minimum"]))
                ok_min = row["Base Diet + Supplements"] + tol >= row["Minimum"]
                ok_max = True if pd.isna(row["Maximum"]) else (
                    row["Base Diet + Supplements"] <= row["Maximum"] + 1e-9 + 1e-6 * max(1.0, abs(row["Maximum"])))
                if ok_min and ok_max:   return "✅ ok"
                if not ok_min:          return "⬇️ below Min"
                if not ok_max:          return "⬆️ above Max"
                return "⚠️ out of range"
            nc_df["Status"] = nc_df.apply(_status, axis=1)
            for c in ["Minimum", "Maximum", "Base Diet", "Supplementation", "Base Diet + Supplements"]:
                nc_df[c] = nc_df[c].round(4)
            nc_df = nc_df.sort_values(["Status", "Nutrient"]).reset_index(drop=True)
            st.dataframe(nc_df, use_container_width=True, hide_index=True,
                         column_config={
                             "Nutrient":                 st.column_config.TextColumn(width="medium"),
                             "Minimum":                  st.column_config.NumberColumn(format="%.2f", width="small"),
                             "Maximum":                  st.column_config.NumberColumn(format="%.2f", width="small"),
                             "Base Diet":                st.column_config.NumberColumn(format="%.2f", width="small"),
                             "Supplementation":          st.column_config.NumberColumn(format="%.2f", width="small"),
                             "Base Diet + Supplements":  st.column_config.NumberColumn(format="%.2f", width="small"),
                             "Status":                   st.column_config.TextColumn(width="small"),
                         })

    # try:
    #     st.image("static/pikachu.jpg", width=230)
    # except Exception:
    #     pass
